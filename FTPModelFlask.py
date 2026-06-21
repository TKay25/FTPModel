from flask import Flask, request, jsonify, send_file, has_request_context
import json
import os
import sqlite3
import gc
import threading
import uuid
import shutil
import pandas as pd
import io
import openpyxl
import numpy as np
from datetime import datetime, timedelta
from time import perf_counter
import time
import re
from reportlab.lib import colors
from reportlab.lib.pagesizes import portrait, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER

app = Flask(__name__)

APP_BASE_DIR = os.path.dirname(__file__)
DATA_ROOT_DIR = os.getenv('FTP_DATA_DIR', APP_BASE_DIR)

CURVE_CONFIG_PATH = os.path.join(DATA_ROOT_DIR, 'curve_config.json')
PROCESSED_REPORTS_DB_PATH = os.path.join(DATA_ROOT_DIR, 'processed_reports.db')
PROCESSED_OUTPUTS_DIR = os.path.join(DATA_ROOT_DIR, 'processed_outputs')
WORKINGS_JSON_DIR = os.path.join(DATA_ROOT_DIR, 'workings_json')
BRANCH_MAP_DB_PATH = os.path.join(DATA_ROOT_DIR, 'branch_sbu_map.db')
BRANCH_MAP_JSON_PATH = os.path.join(DATA_ROOT_DIR, 'branch_sbu_map.json')
UPLOAD_JOBS_DB_PATH = os.path.join(DATA_ROOT_DIR, 'upload_jobs.db')
TEMP_UPLOADS_DIR = os.path.join(DATA_ROOT_DIR, 'temp_uploads')
LOAN_SHEETS = {'ZWG LOANS', 'FX LOANS'}
# Disable non-loan sheet export by default to keep memory use below small-instance limits.
INCLUDE_NON_LOAN_SHEETS = os.getenv('INCLUDE_NON_LOAN_SHEETS', '0').lower() in {'1', 'true', 'yes'}
# Guardrail for low-memory hosts: downgrade full-workbook export if upload is too large.
INCLUDE_WORKINGS_MAX_UPLOAD_MB = float(os.getenv('INCLUDE_WORKINGS_MAX_UPLOAD_MB', '6'))
FORCE_RESULTS_ONLY_ON_HOSTED = os.getenv(
    'FTP_FORCE_RESULTS_ONLY_ON_HOSTED',
    '1' if (os.getenv('RENDER') or os.getenv('RENDER_EXTERNAL_URL')) else '0'
).lower() in {'1', 'true', 'yes', 'on'}
REPORT_RETENTION_MAX_VERSIONS = int(os.getenv('REPORT_RETENTION_MAX_VERSIONS', '3'))
REPORT_RETENTION_MAX_MONTHS = int(os.getenv('REPORT_RETENTION_MAX_MONTHS', '24'))
FTP_REQUIRE_ROLE = os.getenv('FTP_REQUIRE_ROLE', '0').lower() in {'1', 'true', 'yes', 'on'}
FTP_DEFAULT_ROLE = os.getenv('FTP_DEFAULT_ROLE', 'admin').strip().lower() or 'admin'
ENABLE_SCHEDULER = os.getenv('ENABLE_SCHEDULER', '0').lower() in {'1', 'true', 'yes', 'on'}
SCHEDULER_POLL_SECONDS = int(os.getenv('SCHEDULER_POLL_SECONDS', '30'))
ENABLE_NOTIFICATIONS = os.getenv('ENABLE_NOTIFICATIONS', '1').lower() in {'1', 'true', 'yes', 'on'}
os.makedirs(DATA_ROOT_DIR, exist_ok=True)
os.makedirs(PROCESSED_OUTPUTS_DIR, exist_ok=True)
os.makedirs(WORKINGS_JSON_DIR, exist_ok=True)
os.makedirs(TEMP_UPLOADS_DIR, exist_ok=True)
print(f"[INFO] FTP storage root: {DATA_ROOT_DIR}")


DEFAULT_BRANCH_SBU_MAP = {
    '0': {'unit': 'Shared Services', 'sbu': 'Shared Services'},
    '11': {'unit': 'Kwame Nkrumah', 'sbu': 'Retail Banking'},
    '12': {'unit': '8Th Avenue', 'sbu': 'Retail Banking'},
    '13': {'unit': 'Mutare', 'sbu': 'Retail Banking'},
    '14': {'unit': 'Kwekwe', 'sbu': 'Retail Banking'},
    '15': {'unit': 'Chitungwiza', 'sbu': 'Retail Banking'},
    '17': {'unit': 'Gokwe', 'sbu': 'Retail Banking'},
    '18': {'unit': 'Gweru', 'sbu': 'Retail Banking'},
    '20': {'unit': 'Chivhu', 'sbu': 'Retail Banking'},
    '21': {'unit': 'Selous', 'sbu': 'Retail Banking'},
    '23': {'unit': 'Southerton', 'sbu': 'Retail Banking'},
    '24': {'unit': 'Sapphire', 'sbu': 'Retail Banking'},
    '25': {'unit': 'Masvingo', 'sbu': 'Retail Banking'},
    '26': {'unit': 'Belmont', 'sbu': 'Retail Banking'},
    '27': {'unit': 'Cash Depot Bulawayo', 'sbu': 'Retail Banking'},
    '28': {'unit': 'Chiredzi', 'sbu': 'Retail Banking'},
    '29': {'unit': 'Borrowdale', 'sbu': 'Retail Banking'},
    '30': {'unit': 'Avondale', 'sbu': 'Retail Banking'},
    '31': {'unit': 'Chinhoyi', 'sbu': 'Retail Banking'},
    '32': {'unit': 'Kwekwe', 'sbu': 'Retail Banking'},
    '33': {'unit': 'Sapphire', 'sbu': 'Retail Banking'},
    '34': {'unit': 'Cash Depot Harare', 'sbu': 'Retail Banking'},
    '35': {'unit': 'Shared Services', 'sbu': 'Shared Services'},
    '36': {'unit': 'Shared Services', 'sbu': 'Shared Services'},
    '37': {'unit': 'Shared Services', 'sbu': 'Shared Services'},
    '38': {'unit': 'Shared Services', 'sbu': 'Shared Services'},
    '39': {'unit': 'Shared Services', 'sbu': 'Shared Services'},
    '40': {'unit': 'Shared Services', 'sbu': 'Shared Services'},
    '41': {'unit': 'Shared Services', 'sbu': 'Shared Services'},
    '43': {'unit': 'Shared Services', 'sbu': 'Shared Services'},
    '44': {'unit': 'Private Sector', 'sbu': 'Corporate Banking'},
    '45': {'unit': 'Business Banking', 'sbu': 'Corporate Banking'},
    '46': {'unit': 'Shared Services', 'sbu': 'Shared Services'},
    '47': {'unit': 'Private Sector', 'sbu': 'Corporate Banking'},
    '48': {'unit': 'Private Sector', 'sbu': 'Corporate Banking'},
    '49': {'unit': 'Shared Services', 'sbu': 'Shared Services'},
    '50': {'unit': 'Shared Services', 'sbu': 'Shared Services'},
    '51': {'unit': 'Retail Head Office', 'sbu': 'Retail Banking'},
    '52': {'unit': 'Kwame Nkrumah', 'sbu': 'Retail Banking'},
    '53': {'unit': 'Private Sector', 'sbu': 'Corporate Banking'},
    '54': {'unit': 'Shared Services', 'sbu': 'Shared Services'},
    '55': {'unit': 'Shared Services', 'sbu': 'Shared Services'},
    '56': {'unit': 'Shared Services', 'sbu': 'Shared Services'},
    '57': {'unit': 'Shared Services', 'sbu': 'Shared Services'},
    '58': {'unit': 'Shared Services', 'sbu': 'Shared Services'},
    '61': {'unit': 'Shared Services', 'sbu': 'Shared Services'},
    '62': {'unit': '8th Avenue', 'sbu': 'Retail Banking'},
    '65': {'unit': 'Shared Services', 'sbu': 'Shared Services'},
    '66': {'unit': 'Treasury', 'sbu': 'Treasury'},
    '67': {'unit': 'Shared Services', 'sbu': 'Shared Services'},
    '68': {'unit': 'Shared Services', 'sbu': 'Shared Services'},
    '69': {'unit': 'Shared Services', 'sbu': 'Shared Services'},
    '70': {'unit': 'Shared Services', 'sbu': 'Shared Services'},
    '87': {'unit': 'Chisipite', 'sbu': 'Retail Banking'},
    '88': {'unit': '8Th Avenue', 'sbu': 'Retail Banking'},
    '89': {'unit': 'Highfield', 'sbu': 'Retail Banking'},
    '90': {'unit': 'Marondera', 'sbu': 'Retail Banking'},
    '91': {'unit': 'Chitungwiza', 'sbu': 'Retail Banking'},
    '92': {'unit': 'Gokwe', 'sbu': 'Retail Banking'},
    '93': {'unit': 'Beitbridge', 'sbu': 'Retail Banking'},
    '95': {'unit': 'Kariba', 'sbu': 'Retail Banking'},
    '96': {'unit': 'Kariba', 'sbu': 'Retail Banking'},
    '97': {'unit': 'Karoi', 'sbu': 'Retail Banking'},
    '98': {'unit': 'Chinhoyi', 'sbu': 'Retail Banking'},
    '99': {'unit': 'Masvingo', 'sbu': 'Retail Banking'},
    '100': {'unit': 'Mvurwi', 'sbu': 'Retail Banking'},
    '101': {'unit': 'Chipinge', 'sbu': 'Retail Banking'},
    '102': {'unit': 'Rusape', 'sbu': 'Retail Banking'},
    '103': {'unit': 'Murehwa', 'sbu': 'Retail Banking'},
    '104': {'unit': 'Victoria Falls', 'sbu': 'Retail Banking'},
    '105': {'unit': 'Shared Services', 'sbu': 'Shared Services'},
    '106': {'unit': 'Agribusiness', 'sbu': 'Corporate Banking'},
    '107': {'unit': 'Institutional Banking', 'sbu': 'Corporate Banking'},
    '108': {'unit': 'Business Banking', 'sbu': 'Corporate Banking'},
    '109': {'unit': 'Chiredzi', 'sbu': 'Retail Banking'},
    '110': {'unit': 'Selous', 'sbu': 'Retail Banking'},
    '111': {'unit': 'Selous', 'sbu': 'Retail Banking'},
    '112': {'unit': 'Mvurwi', 'sbu': 'Retail Banking'},
    '113': {'unit': 'Custodial Services', 'sbu': 'Corporate Banking'},
    '114': {'unit': 'Retail Head Office', 'sbu': 'Retail Banking'},
    '115': {'unit': 'Shared Services', 'sbu': 'Shared Services'},
    '116': {'unit': 'Shared Services', 'sbu': 'Shared Services'},
    '117': {'unit': 'Shared Services', 'sbu': 'Shared Services'},
    '118': {'unit': 'Bureau De Change Hire', 'sbu': 'Shared Services'},
    '120': {'unit': 'Sapphire', 'sbu': 'Retail Banking'},
    '121': {'unit': 'Retail Centralised', 'sbu': 'Retail Banking'},
    '122': {'unit': 'Mat Centre Fire Street', 'sbu': 'Retail Banking'},
    '123': {'unit': 'Shared Services', 'sbu': 'Shared Services'},
    '124': {'unit': 'Shared Services', 'sbu': 'Shared Services'},
    '125': {'unit': 'Passport Centre Harare', 'sbu': 'Retail Banking'},
    '126': {'unit': 'Virtual Branch', 'sbu': 'Retail Banking'},
    '127': {'unit': 'Passport Centre Harare', 'sbu': 'Retail Banking'},
    '128': {'unit': 'Passport Centre Chitungwiza', 'sbu': 'Retail Banking'},
    '129': {'unit': 'Passport Centre Lupar', 'sbu': 'Retail Banking'},
    '130': {'unit': 'Passport Centre Hwange', 'sbu': 'Retail Banking'},
    '131': {'unit': 'Passport Centre Gweru', 'sbu': 'Retail Banking'},
    '132': {'unit': 'Passport Centre Beitbridge', 'sbu': 'Retail Banking'},
    '133': {'unit': 'Passport Centre Chinhoyi', 'sbu': 'Retail Banking'},
    '134': {'unit': 'Passport Centre Marondera', 'sbu': 'Retail Banking'},
    '135': {'unit': 'Passport Centre Bindura', 'sbu': 'Retail Banking'},
    '136': {'unit': 'Passport Centre Gwe', 'sbu': 'Retail Banking'},
    '137': {'unit': 'Passport Centre Mutare', 'sbu': 'Retail Banking'},
    '138': {'unit': 'Passport Centre Mvurwi', 'sbu': 'Retail Banking'},
    '139': {'unit': 'Passport Centre Zvishavane', 'sbu': 'Retail Banking'},
    '140': {'unit': 'Passport Centre Murehwa', 'sbu': 'Retail Banking'},
    '141': {'unit': 'Shared Services', 'sbu': 'Shared Services'},
    '142': {'unit': 'Retail Centralised', 'sbu': 'Retail Banking'},
    '143': {'unit': 'Retail Head Office', 'sbu': 'Retail Banking'},
    '144': {'unit': 'Retail Head Office', 'sbu': 'Retail Banking'},
    '145': {'unit': 'Borrowdale', 'sbu': 'Retail Banking'},
    '146': {'unit': 'Passport Centre Mwer', 'sbu': 'Retail Banking'},
    '147': {'unit': 'Passport Centre Gokwe', 'sbu': 'Retail Banking'},
    '203': {'unit': 'Shared Services', 'sbu': 'Shared Services'},
    '600': {'unit': 'Shared Services', 'sbu': 'Shared Services'},
    '601': {'unit': 'Treasury', 'sbu': 'Treasury'},
    '602': {'unit': 'Mortgage Finance', 'sbu': 'Retail Banking'},
    '611': {'unit': 'Masvingo', 'sbu': 'Retail Banking'},
    '612': {'unit': 'Chiredzi', 'sbu': 'Retail Banking'},
    '613': {'unit': 'Chinhoyi', 'sbu': 'Retail Banking'},
    '614': {'unit': 'Zvishavane', 'sbu': 'Retail Banking'},
    '615': {'unit': 'Gweru', 'sbu': 'Retail Banking'},
    '616': {'unit': 'Kwekwe', 'sbu': 'Retail Banking'},
    '617': {'unit': 'Kadoma', 'sbu': 'Retail Banking'},
    '618': {'unit': 'Kadoma', 'sbu': 'Retail Banking'},
    '619': {'unit': 'Gokwe', 'sbu': 'Retail Banking'},
    '629': {'unit': 'Chipinge', 'sbu': 'Retail Banking'},
    '630': {'unit': 'Chipinge', 'sbu': 'Retail Banking'},
    '631': {'unit': 'Mutare', 'sbu': 'Retail Banking'},
    '632': {'unit': 'Mutare', 'sbu': 'Retail Banking'},
    '633': {'unit': 'Mutare', 'sbu': 'Retail Banking'},
    '634': {'unit': 'Rusape', 'sbu': 'Retail Banking'},
    '644': {'unit': '8th Avenue', 'sbu': 'Retail Banking'},
    '645': {'unit': '8th Avenue', 'sbu': 'Retail Banking'},
    '646': {'unit': 'Belmont', 'sbu': 'Retail Banking'},
    '647': {'unit': 'Belmont', 'sbu': 'Retail Banking'},
    '648': {'unit': 'Belmont', 'sbu': 'Retail Banking'},
    '649': {'unit': 'Gwanda', 'sbu': 'Retail Banking'},
    '658': {'unit': 'Cash Depot Bulawayo', 'sbu': 'Retail Banking'},
    '660': {'unit': 'Samora Machel', 'sbu': 'Retail Banking'},
    '661': {'unit': 'Avondale', 'sbu': 'Retail Banking'},
    '662': {'unit': 'Bindura', 'sbu': 'Retail Banking'},
    '663': {'unit': 'Msasa', 'sbu': 'Retail Banking'},
    '664': {'unit': 'Chinhoyi', 'sbu': 'Retail Banking'},
    '665': {'unit': 'Sapphire', 'sbu': 'Retail Banking'},
    '667': {'unit': 'Karoi', 'sbu': 'Retail Banking'},
    '668': {'unit': 'Murehwa', 'sbu': 'Retail Banking'},
    '669': {'unit': 'Samora Machel', 'sbu': 'Retail Banking'},
    '670': {'unit': 'Samora Machel', 'sbu': 'Retail Banking'},
    '671': {'unit': 'Cash Depot Harare', 'sbu': 'Retail Banking'},
    '672': {'unit': 'Kariba', 'sbu': 'Retail Banking'},
    '681': {'unit': 'Sapphire', 'sbu': 'Retail Banking'},
    '682': {'unit': 'Chivhu', 'sbu': 'Retail Banking'},
    '683': {'unit': 'Chitungwiza', 'sbu': 'Retail Banking'},
    '684': {'unit': 'Chivhu', 'sbu': 'Retail Banking'},
    '685': {'unit': 'Sapphire', 'sbu': 'Retail Banking'},
    '686': {'unit': 'Highfield', 'sbu': 'Retail Banking'},
    '687': {'unit': 'Marondera', 'sbu': 'Retail Banking'},
    '688': {'unit': 'Msasa', 'sbu': 'Retail Banking'},
    '689': {'unit': 'Msasa', 'sbu': 'Retail Banking'},
    '690': {'unit': 'Sapphire', 'sbu': 'Retail Banking'}
}


def init_branch_map_db():
    if os.path.exists(BRANCH_MAP_JSON_PATH):
        return

    migrated_map = None
    if os.path.exists(BRANCH_MAP_DB_PATH):
        try:
            with sqlite3.connect(BRANCH_MAP_DB_PATH) as connection:
                rows = connection.execute(
                    'SELECT branch_code, unit, sbu FROM branch_sbu_map'
                ).fetchall()
            if rows:
                migrated_map = {
                    str(code): {'unit': str(unit), 'sbu': str(sbu)}
                    for code, unit, sbu in rows
                }
        except sqlite3.Error:
            migrated_map = None

    if not migrated_map:
        migrated_map = DEFAULT_BRANCH_SBU_MAP

    with open(BRANCH_MAP_JSON_PATH, 'w', encoding='utf-8') as map_file:
        json.dump(migrated_map, map_file, indent=2, ensure_ascii=True)


def load_branch_sbu_map():
    init_branch_map_db()
    try:
        with open(BRANCH_MAP_JSON_PATH, 'r', encoding='utf-8') as map_file:
            stored_map = json.load(map_file)
        if not isinstance(stored_map, dict):
            raise ValueError('Invalid branch map format')

        normalized_map = {}
        for code, entry in stored_map.items():
            if not isinstance(entry, dict):
                continue
            unit = str(entry.get('unit', '')).strip()
            sbu = str(entry.get('sbu', '')).strip()
            code_text = str(code).strip()
            if not code_text or not unit or not sbu:
                continue
            normalized_map[code_text] = {'unit': unit, 'sbu': sbu}

        if not normalized_map:
            raise ValueError('Branch map is empty')
        return normalized_map
    except (OSError, ValueError, TypeError, json.JSONDecodeError):
        with open(BRANCH_MAP_JSON_PATH, 'w', encoding='utf-8') as map_file:
            json.dump(DEFAULT_BRANCH_SBU_MAP, map_file, indent=2, ensure_ascii=True)
        return DEFAULT_BRANCH_SBU_MAP.copy()


def load_branch_sbu_rows():
    branch_map = load_branch_sbu_map()

    def _sort_key(code_text):
        try:
            return (0, int(code_text), code_text)
        except ValueError:
            return (1, code_text)

    rows = []
    for code in sorted(branch_map.keys(), key=_sort_key):
        entry = branch_map[code]
        rows.append({'code': str(code), 'unit': entry['unit'], 'sbu': entry['sbu']})
    return rows


def save_branch_sbu_rows(rows):
    normalized_rows = []
    seen_codes = set()

    for row in rows:
        code = str(row.get('code', '')).strip()
        unit = str(row.get('unit', '')).strip()
        sbu = str(row.get('sbu', '')).strip()

        if not code or not unit or not sbu:
            continue

        if code in seen_codes:
            continue

        seen_codes.add(code)
        normalized_rows.append((code, unit, sbu))

    if not normalized_rows:
        raise ValueError('At least one valid branch mapping is required')

    payload = {
        code: {'unit': unit, 'sbu': sbu}
        for code, unit, sbu in normalized_rows
    }
    with open(BRANCH_MAP_JSON_PATH, 'w', encoding='utf-8') as map_file:
        json.dump(payload, map_file, indent=2, ensure_ascii=True)


def reset_branch_sbu_map_to_default():
    save_branch_sbu_rows([
        {'code': code, 'unit': data['unit'], 'sbu': data['sbu']}
        for code, data in DEFAULT_BRANCH_SBU_MAP.items()
    ])


init_branch_map_db()


def load_curve_config():
    default_curves = {
        'tenors': [7, 14, 21, 30, 60, 90, 180, 270, 360, 720, 1080, 1460, 1800],
        'usd_rates': [3.29, 3.36, 6.15, 10.97, 11.02, 11.13, 12.22, 12.22, 12.22, 13.96, 15.41, 18.32, 18.32],
        'zwg_rates': [16.90, 16.90, 16.90, 16.90, 17.90, 18.10, 19.10, 19.10, 20.10, 23.47, 26.54, 32.67, 32.67]
    }

    if not os.path.exists(CURVE_CONFIG_PATH):
        return default_curves

    try:
        with open(CURVE_CONFIG_PATH, 'r', encoding='utf-8') as curve_file:
            curve_config = json.load(curve_file)

        tenors = [int(value) for value in curve_config.get('tenors', default_curves['tenors'])]
        usd_curve = [float(value) for value in curve_config.get('usd_rates', default_curves['usd_rates'])]
        zwg_curve = [float(value) for value in curve_config.get('zwg_rates', default_curves['zwg_rates'])]

        if len(tenors) != len(usd_curve) or len(tenors) != len(zwg_curve):
            raise ValueError('Curve config lengths do not match')

        return {
            'tenors': tenors,
            'usd_rates': usd_curve,
            'zwg_rates': zwg_curve
        }
    except (OSError, ValueError, TypeError, json.JSONDecodeError):
        return default_curves


def save_curve_config(curve_config):
    with open(CURVE_CONFIG_PATH, 'w', encoding='utf-8') as curve_file:
        json.dump(curve_config, curve_file, indent=2)

curve_config = load_curve_config()

MONTH_NAME_TO_NUMBER = {
    'january': 1,
    'february': 2,
    'march': 3,
    'april': 4,
    'may': 5,
    'june': 6,
    'july': 7,
    'august': 8,
    'september': 9,
    'october': 10,
    'november': 11,
    'december': 12,
}
MONTH_NUMBER_TO_NAME = {value: name.title() for name, value in MONTH_NAME_TO_NUMBER.items()}

# Define the tenor points (in days)
tenors = curve_config['tenors']

# USD FTP curve data
usd_rates = curve_config['usd_rates']

# ZWG FTP curve data
zwg_rates = curve_config['zwg_rates']

# Global variable
latest_data = {
    'filename': None,
    'sheets': {},
    'ftp_results': None,
    'summaries': {},
    'period': {},
    'excel_output_path': None,
    'excel_filename': None,
    'original_upload_path': None,
    'original_upload_filename': None,
    'workings_manifest_path': None,
    'excel_contains_workings': False,
    'skipped_sheet_count': 0,
    'export_warning': None,
}


def _json_default(value):
    if isinstance(value, np.generic):
        return value.item()
    if isinstance(value, (datetime, timedelta)):
        return str(value)
    return str(value)


def _sanitize_json_value(value):
    if isinstance(value, np.generic):
        return _sanitize_json_value(value.item())
    if isinstance(value, dict):
        return {key: _sanitize_json_value(val) for key, val in value.items()}
    if isinstance(value, list):
        return [_sanitize_json_value(item) for item in value]
    if isinstance(value, float):
        if np.isnan(value) or np.isinf(value):
            return None
        return value
    if isinstance(value, (datetime, timedelta)):
        return str(value)
    if isinstance(value, pd.Timestamp):
        if pd.isna(value):
            return None
        return value.isoformat()
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    return value


def _strict_json_response(payload, status_code=200):
    sanitized_payload = _sanitize_json_value(payload)
    response_body = json.dumps(sanitized_payload, ensure_ascii=True, default=_json_default, allow_nan=False)
    return app.response_class(response=response_body, status=status_code, mimetype='application/json')


def _safe_sheet_token(sheet_name):
    return re.sub(r'[^A-Za-z0-9]+', '_', str(sheet_name)).strip('_').lower() or 'sheet'


def _build_workings_manifest_path(report_key):
    return os.path.join(WORKINGS_JSON_DIR, f'FTP_Workings_{report_key}.json')


def _build_workings_sheet_data_path(report_key, sheet_name):
    token = _safe_sheet_token(sheet_name)
    return os.path.join(WORKINGS_JSON_DIR, f'FTP_Workings_{report_key}_{token}.ndjson')


def _normalize_sheet_columns(raw_columns):
    seen = {}
    normalized = []
    for index, value in enumerate(raw_columns or []):
        base_name = str(value).strip() if value is not None and str(value).strip() else f'Column_{index + 1}'
        duplicate_count = seen.get(base_name, 0)
        normalized_name = base_name if duplicate_count == 0 else f'{base_name}.{duplicate_count}'
        seen[base_name] = duplicate_count + 1
        normalized.append(normalized_name)
    return normalized


def _stream_workbook_sheet_to_ndjson(workbook_path, sheet_name, output_path):
    extension = os.path.splitext(str(workbook_path))[1].lower()
    if extension not in {'.xlsx', '.xlsm', '.xltx', '.xltm'}:
        return None

    workbook = None
    try:
        workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=False)
        if sheet_name not in workbook.sheetnames:
            return None

        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if header_row is None:
            with open(output_path, 'w', encoding='utf-8'):
                pass
            return {'columns': [], 'row_count': 0}

        columns = _normalize_sheet_columns(header_row)
        row_count = 0
        with open(output_path, 'w', encoding='utf-8') as output_file:
            for row_values in rows:
                row_payload = {
                    column: _sanitize_json_value(row_values[index] if row_values and index < len(row_values) else None)
                    for index, column in enumerate(columns)
                }
                output_file.write(json.dumps(row_payload, ensure_ascii=True, allow_nan=False, default=_json_default))
                output_file.write('\n')
                row_count += 1

        return {'columns': columns, 'row_count': row_count}
    except Exception as exc:
        print(f"[WARN] Failed to stream {sheet_name} directly from workbook: {exc}")
        if os.path.exists(output_path):
            try:
                os.remove(output_path)
            except OSError:
                pass
        return None
    finally:
        if workbook is not None:
            workbook.close()


def _write_dataframe_ndjson(df, output_path):
    columns = list(df.columns)
    with open(output_path, 'w', encoding='utf-8') as output_file:
        for row_values in df.itertuples(index=False, name=None):
            row_payload = {
                column: _sanitize_json_value(value)
                for column, value in zip(columns, row_values)
            }
            output_file.write(json.dumps(row_payload, ensure_ascii=True, allow_nan=False, default=_json_default))
            output_file.write('\n')


def _delete_workings_artifacts(report_key):
    if not report_key:
        return

    manifest_path = _build_workings_manifest_path(report_key)
    if os.path.exists(manifest_path):
        try:
            with open(manifest_path, 'r', encoding='utf-8') as manifest_file:
                manifest_payload = json.load(manifest_file)
            for sheet_info in (manifest_payload.get('sheets') or {}).values():
                data_path = sheet_info.get('data_path')
                if data_path and os.path.exists(data_path):
                    os.remove(data_path)
        except Exception as exc:
            print(f"[WARN] Failed to remove workings data files for {report_key}: {exc}")
        finally:
            try:
                os.remove(manifest_path)
            except OSError:
                pass


def _ensure_processed_reports_db():
    with sqlite3.connect(PROCESSED_REPORTS_DB_PATH) as connection:
        connection.execute(
            '''
            CREATE TABLE IF NOT EXISTS processed_reports (
                report_key TEXT PRIMARY KEY,
                month_number INTEGER NOT NULL,
                month_name TEXT NOT NULL,
                year INTEGER NOT NULL,
                filename TEXT,
                period_json TEXT NOT NULL,
                summaries_json TEXT NOT NULL,
                sheets_json TEXT NOT NULL,
                excel_output_path TEXT,
                excel_filename TEXT,
                original_upload_path TEXT,
                original_upload_filename TEXT,
                excel_contains_workings INTEGER NOT NULL DEFAULT 0,
                skipped_sheet_count INTEGER NOT NULL DEFAULT 0,
                export_warning TEXT,
                created_at REAL NOT NULL,
                updated_at REAL NOT NULL
            )
            '''
        )
        existing_columns = {
            row[1] for row in connection.execute('PRAGMA table_info(processed_reports)').fetchall()
        }
        if 'original_upload_path' not in existing_columns:
            connection.execute('ALTER TABLE processed_reports ADD COLUMN original_upload_path TEXT')
        if 'original_upload_filename' not in existing_columns:
            connection.execute('ALTER TABLE processed_reports ADD COLUMN original_upload_filename TEXT')
        connection.execute(
            '''
            CREATE TABLE IF NOT EXISTS report_notes (
                note_id INTEGER PRIMARY KEY AUTOINCREMENT,
                report_key TEXT NOT NULL,
                note_text TEXT NOT NULL,
                created_by TEXT,
                created_at REAL NOT NULL
            )
            '''
        )
        connection.execute(
            '''
            CREATE TABLE IF NOT EXISTS audit_log (
                event_id INTEGER PRIMARY KEY AUTOINCREMENT,
                event_type TEXT NOT NULL,
                report_key TEXT,
                actor TEXT,
                actor_role TEXT,
                details_json TEXT,
                created_at REAL NOT NULL
            )
            '''
        )
        connection.execute(
            '''
            CREATE TABLE IF NOT EXISTS scheduler_jobs (
                job_id INTEGER PRIMARY KEY AUTOINCREMENT,
                folder_path TEXT NOT NULL,
                include_workings INTEGER NOT NULL DEFAULT 0,
                overwrite_existing INTEGER NOT NULL DEFAULT 0,
                active INTEGER NOT NULL DEFAULT 1,
                created_at REAL NOT NULL,
                updated_at REAL NOT NULL,
                last_run_at REAL,
                last_status TEXT,
                last_message TEXT
            )
            '''
        )
        connection.commit()


def _current_actor():
    if not has_request_context():
        return 'system'
    actor = request.headers.get('X-FTP-Actor') or request.args.get('actor')
    if not actor:
        actor = 'local-user'
    return str(actor).strip() or 'local-user'


def _current_role():
    if not has_request_context():
        return 'system'
    role = request.headers.get('X-FTP-Role') or request.args.get('role')
    if not role:
        role = FTP_DEFAULT_ROLE
    return str(role).strip().lower() or FTP_DEFAULT_ROLE


def _require_role(*allowed_roles):
    if not FTP_REQUIRE_ROLE:
        return None

    role = _current_role()
    normalized_allowed = {str(value).strip().lower() for value in allowed_roles}
    if role not in normalized_allowed:
        return jsonify({'error': f'Action requires one of the following roles: {", ".join(sorted(normalized_allowed))}'}), 403
    return None


def _audit_event(event_type, report_key=None, details=None):
    _ensure_processed_reports_db()
    with sqlite3.connect(PROCESSED_REPORTS_DB_PATH) as connection:
        connection.execute(
            '''
            INSERT INTO audit_log (event_type, report_key, actor, actor_role, details_json, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
            ''',
            (
                event_type,
                report_key,
                _current_actor(),
                _current_role(),
                json.dumps(details or {}, ensure_ascii=True, default=_json_default),
                time.time(),
            )
        )
        connection.commit()


def _register_notification(message, level='info', report_key=None):
    if not ENABLE_NOTIFICATIONS:
        return
    _audit_event('notification', report_key=report_key, details={'message': message, 'level': level})


def _fetch_report_metadata(report_key):
    _ensure_processed_reports_db()
    with sqlite3.connect(PROCESSED_REPORTS_DB_PATH) as connection:
        connection.row_factory = sqlite3.Row
        cursor = connection.cursor()
        cursor.execute(
            '''
            SELECT report_key, month_number, month_name, year, filename, excel_filename,
                   excel_output_path, original_upload_path, original_upload_filename,
                   excel_contains_workings, skipped_sheet_count, export_warning,
                   created_at, updated_at, period_json, summaries_json, sheets_json
            FROM processed_reports
            WHERE report_key = ?
            ''',
            (report_key,)
        )
        return cursor.fetchone()


def _load_report_notes(report_key):
    _ensure_processed_reports_db()
    with sqlite3.connect(PROCESSED_REPORTS_DB_PATH) as connection:
        connection.row_factory = sqlite3.Row
        rows = connection.execute(
            'SELECT note_id, note_text, created_by, created_at FROM report_notes WHERE report_key = ? ORDER BY created_at DESC',
            (report_key,)
        ).fetchall()
    return [dict(row) for row in rows]


def _delete_report_workbook(excel_output_path):
    if excel_output_path and os.path.exists(excel_output_path):
        try:
            os.remove(excel_output_path)
        except OSError as exc:
            print(f"[WARN] Failed to remove archived workbook {excel_output_path}: {exc}")


def _delete_original_upload(original_upload_path):
    if original_upload_path and os.path.exists(original_upload_path):
        try:
            os.remove(original_upload_path)
        except OSError as exc:
            print(f"[WARN] Failed to remove archived original upload {original_upload_path}: {exc}")


def _cleanup_report_retention(month_number=None, year=None):
    _ensure_processed_reports_db()
    now = datetime.now()
    cutoff_year = now.year
    cutoff_month = now.month - REPORT_RETENTION_MAX_MONTHS
    while cutoff_month <= 0:
        cutoff_month += 12
        cutoff_year -= 1

    with sqlite3.connect(PROCESSED_REPORTS_DB_PATH) as connection:
        connection.row_factory = sqlite3.Row
        rows_to_delete = []

        if REPORT_RETENTION_MAX_MONTHS > 0:
            old_rows = connection.execute(
                '''
                SELECT report_key, excel_output_path, original_upload_path
                FROM processed_reports
                WHERE (year < ?) OR (year = ? AND month_number < ?)
                ''',
                (cutoff_year, cutoff_year, cutoff_month)
            ).fetchall()
            rows_to_delete.extend(old_rows)

        if month_number and year and REPORT_RETENTION_MAX_VERSIONS > 0:
            version_rows = connection.execute(
                '''
                SELECT report_key, excel_output_path, original_upload_path
                FROM processed_reports
                WHERE month_number = ? AND year = ?
                ORDER BY updated_at DESC
                ''',
                (int(month_number), int(year))
            ).fetchall()
            rows_to_delete.extend(version_rows[REPORT_RETENTION_MAX_VERSIONS:])

        unique_rows = {}
        for row in rows_to_delete:
            unique_rows[row['report_key']] = {
                'excel_output_path': row['excel_output_path'],
                'original_upload_path': row['original_upload_path'],
            }

        if not unique_rows:
            return 0

        for report_key, output_paths in unique_rows.items():
            connection.execute('DELETE FROM report_notes WHERE report_key = ?', (report_key,))
            connection.execute('DELETE FROM processed_reports WHERE report_key = ?', (report_key,))
            _audit_event('retention_delete', report_key=report_key, details={'reason': 'retention_policy'})
            _delete_report_workbook(output_paths.get('excel_output_path'))
            _delete_original_upload(output_paths.get('original_upload_path'))
            _delete_workings_artifacts(report_key)

        connection.commit()
        return len(unique_rows)


def save_latest_data_snapshot():
    period = latest_data.get('period') or {}
    month_number = period.get('month_number')
    if not month_number:
        month_value = str(period.get('month', '')).strip().lower()
        month_number = MONTH_NAME_TO_NUMBER.get(month_value)

    year_value = period.get('year')
    if not month_number or year_value is None:
        return

    month_number = int(month_number)
    year_value = int(year_value)
    report_key = period.get('report_key') or f'{year_value:04d}-{month_number:02d}'
    payload = {
        'filename': latest_data.get('filename'),
        'sheets': latest_data.get('sheets', {}),
        'ftp_results': latest_data.get('ftp_results'),
        'summaries': latest_data.get('summaries', {}),
        'period': {
            **period,
            'month_number': month_number,
            'report_key': report_key,
        },
        'excel_output_path': latest_data.get('excel_output_path'),
        'excel_filename': latest_data.get('excel_filename'),
        'original_upload_path': latest_data.get('original_upload_path'),
        'original_upload_filename': latest_data.get('original_upload_filename'),
        'excel_contains_workings': bool(latest_data.get('excel_contains_workings')),
        'skipped_sheet_count': int(latest_data.get('skipped_sheet_count') or 0),
        'export_warning': latest_data.get('export_warning'),
    }

    _ensure_processed_reports_db()
    now_ts = time.time()
    with sqlite3.connect(PROCESSED_REPORTS_DB_PATH) as connection:
        connection.execute(
            '''
            INSERT INTO processed_reports (
                report_key, month_number, month_name, year, filename,
                period_json, summaries_json, sheets_json,
                excel_output_path, excel_filename,
                original_upload_path, original_upload_filename,
                excel_contains_workings, skipped_sheet_count, export_warning,
                created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(report_key) DO UPDATE SET
                month_number = excluded.month_number,
                month_name = excluded.month_name,
                year = excluded.year,
                filename = excluded.filename,
                period_json = excluded.period_json,
                summaries_json = excluded.summaries_json,
                sheets_json = excluded.sheets_json,
                excel_output_path = excluded.excel_output_path,
                excel_filename = excluded.excel_filename,
                original_upload_path = excluded.original_upload_path,
                original_upload_filename = excluded.original_upload_filename,
                excel_contains_workings = excluded.excel_contains_workings,
                skipped_sheet_count = excluded.skipped_sheet_count,
                export_warning = excluded.export_warning,
                updated_at = excluded.updated_at
            ''',
            (
                report_key,
                month_number,
                MONTH_NUMBER_TO_NAME[month_number],
                year_value,
                payload['filename'],
                json.dumps(payload['period'], ensure_ascii=True, default=_json_default),
                json.dumps(payload['summaries'], ensure_ascii=True, default=_json_default),
                json.dumps(payload['sheets'], ensure_ascii=True, default=_json_default),
                payload['excel_output_path'],
                payload['excel_filename'],
                payload['original_upload_path'],
                payload['original_upload_filename'],
                int(payload['excel_contains_workings']),
                int(payload['skipped_sheet_count']),
                payload['export_warning'],
                now_ts,
                now_ts,
            )
        )
        connection.commit()

    _audit_event(
        'report_saved',
        report_key=report_key,
        details={
            'month': payload['period'].get('month'),
            'year': payload['period'].get('year'),
            'excel_contains_workings': payload['excel_contains_workings'],
            'skipped_sheet_count': payload['skipped_sheet_count'],
            'original_upload_available': bool(payload['original_upload_path']),
            'filename': payload['filename'],
        }
    )
    _register_notification(f'Report {report_key} is ready for download.', level='success', report_key=report_key)
    _cleanup_report_retention(month_number=month_number, year=year_value)


def _normalize_month_number(month_value):
    if month_value is None:
        return None
    month_text = str(month_value).strip().lower()
    if not month_text:
        return None
    if month_text.isdigit():
        month_number = int(month_text)
    else:
        month_number = MONTH_NAME_TO_NUMBER.get(month_text)
    if month_number is None or month_number < 1 or month_number > 12:
        return None
    return month_number


def _make_report_key(year_value, month_number, version_suffix=None):
    base_key = f'{int(year_value):04d}-{int(month_number):02d}'
    if version_suffix is None:
        return base_key
    return f'{base_key}-{int(version_suffix)}'


def _load_processed_report_row(month=None, year=None, report_key=None):
    _ensure_processed_reports_db()
    with sqlite3.connect(PROCESSED_REPORTS_DB_PATH) as connection:
        connection.row_factory = sqlite3.Row
        cursor = connection.cursor()

        if report_key:
            cursor.execute('SELECT * FROM processed_reports WHERE report_key = ?', (report_key,))
            return cursor.fetchone()

        if month is None and year is None:
            cursor.execute('SELECT * FROM processed_reports ORDER BY updated_at DESC LIMIT 1')
            return cursor.fetchone()

        month_number = _normalize_month_number(month)
        if not month_number or year is None:
            return None

        cursor.execute(
            'SELECT * FROM processed_reports WHERE month_number = ? AND year = ? ORDER BY updated_at DESC LIMIT 1',
            (int(month_number), int(year))
        )
        return cursor.fetchone()


def _next_report_version_suffix(year_value, month_number):
    base_key = _make_report_key(year_value, month_number)
    with sqlite3.connect(PROCESSED_REPORTS_DB_PATH) as connection:
        cursor = connection.cursor()
        cursor.execute(
            'SELECT report_key FROM processed_reports WHERE report_key = ? OR report_key LIKE ? ORDER BY report_key',
            (base_key, f'{base_key}-%')
        )
        existing_keys = [row[0] for row in cursor.fetchall()]

    if base_key not in existing_keys:
        return None

    suffixes = []
    for existing_key in existing_keys:
        if existing_key == base_key:
            suffixes.append(0)
            continue
        try:
            suffixes.append(int(existing_key.rsplit('-', 1)[1]))
        except (IndexError, ValueError):
            continue

    suffixes = sorted(set(suffixes))
    next_suffix = 1
    for suffix in suffixes:
        if suffix == next_suffix:
            next_suffix += 1
    return next_suffix


def load_latest_data_snapshot(month=None, year=None, report_key=None):
    global latest_data

    try:
        snapshot = _load_processed_report_row(month=month, year=year, report_key=report_key)
        if not snapshot:
            return False

        period = json.loads(snapshot['period_json']) if snapshot['period_json'] else {}
        latest_data = {
            'filename': snapshot['filename'],
            'sheets': json.loads(snapshot['sheets_json']) if snapshot['sheets_json'] else {},
            'ftp_results': None,
            'summaries': json.loads(snapshot['summaries_json']) if snapshot['summaries_json'] else {},
            'period': period,
            'excel_output_path': snapshot['excel_output_path'],
            'excel_filename': snapshot['excel_filename'],
            'original_upload_path': snapshot['original_upload_path'],
            'original_upload_filename': snapshot['original_upload_filename'],
            'workings_manifest_path': _build_workings_manifest_path(period.get('report_key') or snapshot['report_key']),
            'excel_contains_workings': bool(snapshot['excel_contains_workings']),
            'skipped_sheet_count': int(snapshot['skipped_sheet_count'] or 0),
            'export_warning': snapshot['export_warning'],
        }
        return True
    except (OSError, ValueError, TypeError, json.JSONDecodeError, sqlite3.Error, KeyError):
        return False


def delete_processed_report(report_key):
    _ensure_processed_reports_db()
    with sqlite3.connect(PROCESSED_REPORTS_DB_PATH) as connection:
        connection.row_factory = sqlite3.Row
        cursor = connection.cursor()
        cursor.execute('SELECT excel_output_path, original_upload_path FROM processed_reports WHERE report_key = ?', (report_key,))
        row = cursor.fetchone()
        if not row:
            return False

        excel_output_path = row['excel_output_path']
        original_upload_path = row['original_upload_path']
        cursor.execute('DELETE FROM report_notes WHERE report_key = ?', (report_key,))
        cursor.execute('DELETE FROM processed_reports WHERE report_key = ?', (report_key,))
        connection.commit()

    _delete_report_workbook(excel_output_path)
    _delete_original_upload(original_upload_path)
    _delete_workings_artifacts(report_key)
    _audit_event('report_deleted', report_key=report_key, details={'deleted_via': 'api'})
    _register_notification(f'Report {report_key} was deleted.', level='warning', report_key=report_key)

    return True


def ensure_latest_data_available(month=None, year=None):
    if latest_data.get('summaries') and latest_data.get('period'):
        return True
    return load_latest_data_snapshot(month=month, year=year)


def _resolve_month_year_args(month_value, year_value):
    if year_value is None:
        return None, None
    month_value = str(month_value).strip().lower() if month_value is not None else ''
    if not month_value:
        return None, None
    if month_value.isdigit():
        month_number = int(month_value)
    else:
        month_number = MONTH_NAME_TO_NUMBER.get(month_value)
    if not month_number:
        return None, None
    return MONTH_NUMBER_TO_NAME[month_number], int(year_value)


def format_number(value):
    try:
        numeric_value = float(value)
    except (TypeError, ValueError):
        return str(value) if value is not None else ''
    return f'{numeric_value:,.2f}'

def generate_pdf_report():
    """Generate PDF report from the preview data and summaries"""
    buffer = io.BytesIO()
    
    global latest_data
    
    if not latest_data.get('summaries') or not latest_data.get('sheets'):
        doc = SimpleDocTemplate(buffer, pagesize=portrait(A4))
        styles = getSampleStyleSheet()
        story = []
        story.append(Paragraph("No data available", styles['Title']))
        story.append(Paragraph("Please upload a file first.", styles['Normal']))
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    doc = SimpleDocTemplate(buffer, pagesize=portrait(A4),
                           rightMargin=72, leftMargin=72,
                           topMargin=72, bottomMargin=72)
    
    styles = getSampleStyleSheet()
    story = []
    
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=24,
                                  textColor=colors.HexColor('#921f1f'), alignment=TA_CENTER, spaceAfter=30)
    header_style = ParagraphStyle('HeaderStyle', parent=styles['Heading2'], fontSize=16,
                                   textColor=colors.HexColor('#632424'), spaceAfter=12, spaceBefore=12)
    
    title = Paragraph("FTP Analysis Report", title_style)
    story.append(title)
    
    period_text = f"{latest_data['period'].get('first_day', 'N/A')} to {latest_data['period'].get('last_day', 'N/A')}"
    period_style = ParagraphStyle('PeriodStyle', parent=styles['Normal'], fontSize=14,
                                   alignment=TA_CENTER, textColor=colors.HexColor('#632424'), spaceAfter=20)
    story.append(Paragraph(period_text, period_style))
    story.append(Spacer(1, 20))
    
    # Summary by Currency
    story.append(Paragraph("FTP Summary by Currency", header_style))
    
    table_data = [['Currency', 'Sheet', 'Total Exposure', 'Total FTP Charge', 'Records']]
    for currency, sheets in latest_data['summaries'].items():
        for sheet_name, sheet_data in sheets.items():
            table_data.append([
                currency, sheet_name,
                format_number(sheet_data['total_exposure']),
                format_number(sheet_data['total_ftp_charge']),
                f"{sheet_data['row_count']:,}"
            ])
    
    table = Table(table_data, colWidths=[1.2*inch, 1.8*inch, 1.5*inch, 1.5*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#921f1f')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    story.append(table)
    story.append(Spacer(1, 20))
    
    # SBU Breakdown - Aggregate by SBU across all sheets per currency
    story.append(Paragraph("SBU Breakdown by Currency", header_style))
    for currency, sheets in latest_data['summaries'].items():
        story.append(Paragraph(f"<b>{currency}</b>", styles['Normal']))
        sbu_aggregates = {}
        
        # Aggregate SBU data from all sheets
        for sheet_data in sheets.values():
            if sheet_data.get('by_sbu'):
                for sbu in sheet_data['by_sbu']:
                    sbu_name = sbu['SBU']
                    if sbu_name not in sbu_aggregates:
                        sbu_aggregates[sbu_name] = {'exposure': 0.0, 'ftp_charge': 0.0}
                    sbu_aggregates[sbu_name]['exposure'] += sbu.get('Currency Exposure + Currency Accrued Reporting', 0)
                    sbu_aggregates[sbu_name]['ftp_charge'] += sbu.get('FTP Charge', 0)
        
        # Build SBU summary table
        sbu_data = [['SBU', 'Exposure', 'FTP Charge']]
        for sbu_name in sorted(sbu_aggregates.keys()):
            sbu_vals = sbu_aggregates[sbu_name]
            sbu_data.append([
                sbu_name,
                format_number(sbu_vals['exposure']),
                format_number(sbu_vals['ftp_charge'])
            ])
        
        if len(sbu_data) > 1:
            sbu_table = Table(sbu_data, colWidths=[2*inch, 2*inch, 2*inch])
            sbu_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#632424')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
            ]))
            story.append(sbu_table)
            story.append(Spacer(1, 15))
    
    # Grand Totals
    total_exposure_zwg = sum(sheet_data['total_exposure'] for currency, sheets in latest_data['summaries'].items() 
                              for sheet_data in sheets.values() if currency == 'ZWG')
    total_ftp_zwg = sum(sheet_data['total_ftp_charge'] for currency, sheets in latest_data['summaries'].items() 
                         for sheet_data in sheets.values() if currency == 'ZWG')
    total_exposure_fx = sum(sheet_data['total_exposure'] for currency, sheets in latest_data['summaries'].items() 
                             for sheet_data in sheets.values() if currency == 'FX')
    total_ftp_fx = sum(sheet_data['total_ftp_charge'] for currency, sheets in latest_data['summaries'].items() 
                        for sheet_data in sheets.values() if currency == 'FX')
    
    story.append(Paragraph("Grand Totals", header_style))
    totals_data = [['Currency', 'Total Exposure', 'Total FTP Charge']]
    if total_exposure_zwg > 0:
        totals_data.append(['ZWG', format_number(total_exposure_zwg), format_number(total_ftp_zwg)])
    if total_exposure_fx > 0:
        totals_data.append(['USD', format_number(total_exposure_fx), format_number(total_ftp_fx)])
    
    totals_table = Table(totals_data, colWidths=[2*inch, 2.5*inch, 2.5*inch])
    totals_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#921f1f')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
    ]))
    story.append(totals_table)
    story.append(Spacer(1, 20))
    
    # Preview Data Section
    story.append(PageBreak())
    story.append(Paragraph("Data Preview (First 10 Rows per Sheet)", header_style))
    
    for sheet_name, sheet_data in latest_data['sheets'].items():
        story.append(Paragraph(f"<b>{sheet_name}</b>", styles['Normal']))
        story.append(Spacer(1, 5))
        
        preview_rows = sheet_data.get('data', [])[:10]
        if preview_rows and sheet_data.get('columns'):
            columns = sheet_data['columns'][:5]
            preview_table_data = [columns[:5]]
            for row in preview_rows:
                row_data = []
                for col in columns[:5]:
                    val = row.get(col, '')
                    if isinstance(val, float):
                        row_data.append(format_number(val))
                    elif len(str(val)) > 20:
                        row_data.append(str(val)[:17] + '...')
                    else:
                        row_data.append(str(val))
                preview_table_data.append(row_data)
            
            col_width = (6.5 * inch) / len(columns[:5])
            preview_table = Table(preview_table_data, colWidths=[col_width] * len(columns[:5]))
            preview_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#632424')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ]))
            story.append(preview_table)
            story.append(Spacer(1, 15))
    
    story.append(Spacer(1, 30))
    footer_text = f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8,
                                   alignment=TA_CENTER, textColor=colors.grey)
    story.append(Paragraph(footer_text, footer_style))
    story.append(Paragraph("FTP Central System", footer_style))
    
    doc.build(story)
    buffer.seek(0)
    return buffer

@app.route('/download-pdf', methods=['GET'])
def download_pdf():
    try:
        global latest_data
        report_key = request.args.get('report_key')
        month = request.args.get('month')
        year = request.args.get('year', type=int)
        if report_key:
            if not load_latest_data_snapshot(report_key=report_key):
                return jsonify({'error': 'No processed data found for that report version.'}), 404
        elif month and year:
            if not load_latest_data_snapshot(month=month, year=year):
                return jsonify({'error': 'No processed data found for that month and year.'}), 404
        elif not ensure_latest_data_available():
            return jsonify({'error': 'No processed data available. Please upload a file first.'}), 404
        pdf_buffer = generate_pdf_report()
        month = latest_data.get('period', {}).get('month', 'Report')
        year = latest_data.get('period', {}).get('year', '')
        filename = f"FTP_Report_{month}_{year}.pdf"
        return send_file(pdf_buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')
    except Exception as e:
        return jsonify({'error': f'Failed to download PDF: {str(e)}'}), 500


@app.route('/download-excel', methods=['GET'])
def download_excel():
    try:
        global latest_data
        report_key = request.args.get('report_key')
        month = request.args.get('month')
        year = request.args.get('year', type=int)
        if report_key:
            if not load_latest_data_snapshot(report_key=report_key):
                return jsonify({'error': 'No processed data found for that report version.'}), 404
        elif month and year:
            if not load_latest_data_snapshot(month=month, year=year):
                return jsonify({'error': 'No processed data found for that month and year.'}), 404
        elif not ensure_latest_data_available():
            return jsonify({'error': 'No processed data available. Please upload a file first.'}), 404

        excel_path = latest_data.get('excel_output_path')
        excel_filename = latest_data.get('excel_filename') or 'FTP_Results.xlsx'

        if not excel_path or not os.path.exists(excel_path):
            return jsonify({'error': 'Processed Excel output is not available. Please upload the file again.'}), 404

        return send_file(
            excel_path,
            as_attachment=True,
            download_name=excel_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'error': f'Failed to download Excel: {str(e)}'}), 500


@app.route('/download-original-workbook', methods=['GET'])
def download_original_workbook():
    try:
        global latest_data
        report_key = request.args.get('report_key')
        month = request.args.get('month')
        year = request.args.get('year', type=int)
        if report_key:
            if not load_latest_data_snapshot(report_key=report_key):
                return jsonify({'error': 'No processed data found for that report version.'}), 404
        elif month and year:
            if not load_latest_data_snapshot(month=month, year=year):
                return jsonify({'error': 'No processed data found for that month and year.'}), 404
        elif not ensure_latest_data_available():
            return jsonify({'error': 'No processed data available. Please upload a file first.'}), 404

        original_path = latest_data.get('original_upload_path')
        original_filename = latest_data.get('original_upload_filename') or 'FTP_Original_Input.xlsx'
        if not original_path or not os.path.exists(original_path):
            return jsonify({'error': 'Original uploaded workbook is not available for this report version.'}), 404

        return send_file(
            original_path,
            as_attachment=True,
            download_name=original_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'error': f'Failed to download original workbook: {str(e)}'}), 500


@app.route('/download-workings-excel', methods=['GET'])
def download_workings_excel():
    try:
        global latest_data
        report_key = request.args.get('report_key')
        month = request.args.get('month')
        year = request.args.get('year', type=int)
        if report_key:
            if not load_latest_data_snapshot(report_key=report_key):
                return jsonify({'error': 'No processed data found for that report version.'}), 404
        elif month and year:
            if not load_latest_data_snapshot(month=month, year=year):
                return jsonify({'error': 'No processed data found for that month and year.'}), 404
        elif not ensure_latest_data_available():
            return jsonify({'error': 'No processed data available. Please upload a file first.'}), 404

        resolved_report_key = latest_data.get('period', {}).get('report_key')
        manifest_path = _build_workings_manifest_path(resolved_report_key)
        if not manifest_path or not os.path.exists(manifest_path):
            return jsonify({'error': 'Workings JSON is not available for this report. Recompute this period to generate workings JSON.'}), 404

        with open(manifest_path, 'r', encoding='utf-8') as manifest_file:
            manifest_payload = json.load(manifest_file)

        sheet_entries = manifest_payload.get('sheets') or {}
        if not sheet_entries:
            return jsonify({'error': 'Workings JSON exists but has no sheet entries.'}), 404

        workbook = openpyxl.Workbook(write_only=True)
        written_sheet_count = 0
        for sheet_name, sheet_info in sheet_entries.items():
            columns = list(sheet_info.get('columns') or [])
            data_path = sheet_info.get('data_path')
            if not columns or not data_path or not os.path.exists(data_path):
                continue

            worksheet = workbook.create_sheet(title=str(sheet_name)[:31])
            written_sheet_count += 1
            worksheet.append(columns)
            with open(data_path, 'r', encoding='utf-8') as rows_file:
                for line in rows_file:
                    if not line.strip():
                        continue
                    row_payload = json.loads(line)
                    worksheet.append([row_payload.get(column) for column in columns])

        if written_sheet_count == 0:
            return jsonify({'error': 'Workings JSON files were not found for this report version.'}), 404

        output_buffer = io.BytesIO()
        workbook.save(output_buffer)
        output_buffer.seek(0)

        period_month = latest_data.get('period', {}).get('month', 'Month')
        period_year = latest_data.get('period', {}).get('year', 'Year')
        download_name = f'FTP_Workings_{period_month}_{period_year}.xlsx'
        return send_file(
            output_buffer,
            as_attachment=True,
            download_name=download_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'error': f'Failed to download workings Excel: {str(e)}'}), 500


@app.route('/download-workings-json', methods=['GET'])
def download_workings_json():
    """Stream workings JSON (NDJSON) as a single downloadable JSON file.
    This is memory-safe because it reads each sheet's NDJSON file
    line-by-line and writes directly to the response, avoiding Excel's
    memory overhead entirely."""
    try:
        global latest_data
        report_key = request.args.get('report_key')
        month = request.args.get('month')
        year = request.args.get('year', type=int)
        if report_key:
            if not load_latest_data_snapshot(report_key=report_key):
                return jsonify({'error': 'No processed data found for that report version.'}), 404
        elif month and year:
            if not load_latest_data_snapshot(month=month, year=year):
                return jsonify({'error': 'No processed data found for that month and year.'}), 404
        elif not ensure_latest_data_available():
            return jsonify({'error': 'No processed data available. Please upload a file first.'}), 404

        resolved_report_key = latest_data.get('period', {}).get('report_key')
        manifest_path = _build_workings_manifest_path(resolved_report_key)
        if not manifest_path or not os.path.exists(manifest_path):
            return jsonify({'error': 'Workings JSON is not available for this report. Recompute this period to generate workings JSON.'}), 404

        with open(manifest_path, 'r', encoding='utf-8') as manifest_file:
            manifest_payload = json.load(manifest_file)

        sheet_entries = manifest_payload.get('sheets') or {}
        if not sheet_entries:
            return jsonify({'error': 'Workings JSON exists but has no sheet entries.'}), 404

        period_month = latest_data.get('period', {}).get('month', 'Month')
        period_year = latest_data.get('period', {}).get('year', 'Year')
        download_name = f'FTP_Workings_{period_month}_{period_year}.json'

        # Build a single JSON payload: { sheet_name: [row, row, ...], ... }
        # Stream each sheet's NDJSON progressively into a buffer.
        output_buffer = io.StringIO()
        output_buffer.write('{\n')
        first_sheet = True
        for sheet_name, sheet_info in sheet_entries.items():
            data_path = sheet_info.get('data_path')
            columns = list(sheet_info.get('columns') or [])
            if not data_path or not os.path.exists(data_path):
                continue

            if not first_sheet:
                output_buffer.write(',\n')
            first_sheet = False
            output_buffer.write(f'  {json.dumps(sheet_name)}: [\n')

            first_row = True
            with open(data_path, 'r', encoding='utf-8') as rows_file:
                for line in rows_file:
                    if not line.strip():
                        continue
                    if not first_row:
                        output_buffer.write(',\n')
                    first_row = False
                    output_buffer.write('    ')
                    output_buffer.write(line.strip())

            output_buffer.write('\n  ]')

        output_buffer.write('\n}\n')
        output_buffer.seek(0)

        return send_file(
            io.BytesIO(output_buffer.getvalue().encode('utf-8')),
            as_attachment=True,
            download_name=download_name,
            mimetype='application/json'
        )
    except Exception as e:
        return jsonify({'error': f'Failed to download workings JSON: {str(e)}'}), 500


def compute_ftp_components(deposit, loan, tenure):
    try:
        d = float(deposit) if deposit else 0
        l = float(loan) if loan else 0
        t = float(tenure) if tenure else 1
        if d <= 0 or l <= 0 or t <= 0:
            return {'charge': '1.12', 'gain': '0.72', 'net': '1.84'}
        risk_factor = min(l / d, 2.0)
        tenure_factor = min(t * 0.07, 1.2)
        charge = round(1.0 + (risk_factor * 0.3) + (tenure_factor * 0.2), 2)
        gain = round(0.6 + (min(d / 200000, 1.5) * 0.3) + (t * 0.04), 2)
        net = round(gain - charge, 2)
        return {'charge': f"{charge:.2f}", 'gain': f"{gain:.2f}", 'net': f"{net:.2f}"}
    except:
        return {'charge': '1.12', 'gain': '0.72', 'net': '1.84'}

# ---------------------------------------------------------------------------
# Async job registry
# ---------------------------------------------------------------------------
_jobs = {}  # job_id -> { status, progress, stage, result, error, created_at, updated_at, completed_at }
_jobs_lock = threading.Lock()
JOB_RETENTION_SECONDS = int(os.environ.get('JOB_RETENTION_SECONDS', '1800'))


def _update_job(job_id, **kwargs):
    now_ts = time.time()

    with _jobs_lock:
        if job_id in _jobs:
            _jobs[job_id].update(kwargs)
            _jobs[job_id]['updated_at'] = now_ts
            if kwargs.get('status') in {'done', 'error'}:
                _jobs[job_id]['completed_at'] = now_ts

    try:
        with sqlite3.connect(UPLOAD_JOBS_DB_PATH) as conn:
            cur = conn.cursor()
            fields = []
            values = []

            if 'status' in kwargs:
                fields.append('status = ?')
                values.append(kwargs['status'])
            if 'progress' in kwargs:
                fields.append('progress = ?')
                values.append(int(kwargs['progress']))
            if 'stage' in kwargs:
                fields.append('stage = ?')
                values.append(kwargs['stage'])
            if 'result' in kwargs:
                fields.append('result_json = ?')
                values.append(json.dumps(kwargs['result']))
            if 'error' in kwargs:
                fields.append('error = ?')
                values.append(kwargs['error'])

            fields.append('updated_at = ?')
            values.append(now_ts)

            if kwargs.get('status') in {'done', 'error'}:
                fields.append('completed_at = ?')
                values.append(now_ts)

            values.append(job_id)
            cur.execute(
                f"UPDATE upload_jobs SET {', '.join(fields)} WHERE job_id = ?",
                values
            )
            conn.commit()
    except Exception as exc:
        print(f"[WARN] Failed to persist job update for {job_id}: {exc}")


def _cleanup_expired_jobs():
    now = time.time()
    with _jobs_lock:
        expired_ids = []
        for jid, job in _jobs.items():
            completed_at = job.get('completed_at')
            if completed_at is None:
                continue
            if now - completed_at > JOB_RETENTION_SECONDS:
                expired_ids.append(jid)
        for jid in expired_ids:
            _jobs.pop(jid, None)

    try:
        threshold = now - JOB_RETENTION_SECONDS
        with sqlite3.connect(UPLOAD_JOBS_DB_PATH) as conn:
            cur = conn.cursor()
            cur.execute(
                "DELETE FROM upload_jobs WHERE completed_at IS NOT NULL AND completed_at < ?",
                (threshold,)
            )
            conn.commit()
    except Exception as exc:
        print(f"[WARN] Failed to cleanup expired upload jobs: {exc}")


def _ensure_upload_jobs_db():
    with sqlite3.connect(UPLOAD_JOBS_DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute('''
            CREATE TABLE IF NOT EXISTS upload_jobs (
                job_id TEXT PRIMARY KEY,
                status TEXT NOT NULL,
                progress INTEGER NOT NULL DEFAULT 0,
                stage TEXT NOT NULL DEFAULT 'Queued',
                result_json TEXT,
                error TEXT,
                created_at REAL NOT NULL,
                updated_at REAL NOT NULL,
                completed_at REAL
            )
        ''')
        conn.commit()


def _insert_upload_job(job_id):
    now_ts = time.time()
    with sqlite3.connect(UPLOAD_JOBS_DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute('''
            INSERT OR REPLACE INTO upload_jobs
            (job_id, status, progress, stage, result_json, error, created_at, updated_at, completed_at)
            VALUES (?, 'running', 0, 'Queued', NULL, NULL, ?, ?, NULL)
        ''', (job_id, now_ts, now_ts))
        conn.commit()


def _fetch_upload_job(job_id):
    with sqlite3.connect(UPLOAD_JOBS_DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute('''
            SELECT status, progress, stage, result_json, error
            FROM upload_jobs
            WHERE job_id = ?
        ''', (job_id,))
        row = cur.fetchone()
        if not row:
            return None
        result_json = row[3]
        result_payload = None
        if result_json:
            try:
                result_payload = json.loads(result_json)
            except Exception:
                result_payload = None
        return {
            'status': row[0],
            'progress': row[1],
            'stage': row[2],
            'result': result_payload,
            'error': row[4]
        }


_ensure_upload_jobs_db()


def _validate_upload_file(file_path, filename):
    validation = {
        'filename_valid': False,
        'required_sheets_present': False,
        'missing_sheets': [],
        'sheet_count': 0,
        'file_size_bytes': os.path.getsize(file_path) if file_path and os.path.exists(file_path) else 0,
        'warnings': [],
    }

    filename_match = re.search(r'FTP Input File (\w+) (\d{4})', filename or '')
    validation['filename_valid'] = bool(filename_match)
    if not validation['filename_valid']:
        validation['warnings'].append('Filename must be: FTP Input File Month Year.xlsx')
        return validation

    try:
        excel_file = pd.ExcelFile(file_path)
        sheet_names = excel_file.sheet_names
    except Exception as exc:
        validation['warnings'].append(f'Could not read workbook: {str(exc)}')
        return validation

    validation['sheet_count'] = len(sheet_names)
    missing_sheets = sorted(LOAN_SHEETS.difference(sheet_names))
    validation['missing_sheets'] = missing_sheets
    validation['required_sheets_present'] = not missing_sheets
    if missing_sheets:
        validation['warnings'].append(f'Missing required sheets: {", ".join(missing_sheets)}')

    if validation['file_size_bytes'] > int(INCLUDE_WORKINGS_MAX_UPLOAD_MB * 1024 * 1024):
        validation['warnings'].append('Workbook size exceeds the safe full-export threshold; results-only mode may be used.')

    return validation


def _merge_processed_loans_into_original_workbook(original_upload_path, processed_loans_path, merged_output_path):
    """Build a full workbook by overlaying processed loan sheets onto a copy of the original upload."""
    shutil.copyfile(original_upload_path, merged_output_path)

    merged_wb = None
    loan_wb = None
    try:
        merged_wb = openpyxl.load_workbook(merged_output_path)
        loan_wb = openpyxl.load_workbook(processed_loans_path, read_only=True, data_only=False)
        original_order = list(merged_wb.sheetnames)

        for sheet_name in loan_wb.sheetnames:
            source_sheet = loan_wb[sheet_name]

            if sheet_name in original_order:
                sheet_index = original_order.index(sheet_name)
            else:
                sheet_index = len(merged_wb.sheetnames)

            if sheet_name in merged_wb.sheetnames:
                merged_wb.remove(merged_wb[sheet_name])

            destination_sheet = merged_wb.create_sheet(title=sheet_name, index=sheet_index)
            for row in source_sheet.iter_rows(values_only=True):
                destination_sheet.append(row)

        merged_wb.save(merged_output_path)
    finally:
        if loan_wb is not None:
            loan_wb.close()
        if merged_wb is not None:
            merged_wb.close()


def _run_scheduler_cycle():
    if not ENABLE_SCHEDULER:
        return 0

    _ensure_processed_reports_db()
    processed_files = 0
    with sqlite3.connect(PROCESSED_REPORTS_DB_PATH) as connection:
        connection.row_factory = sqlite3.Row
        jobs = connection.execute(
            'SELECT job_id, folder_path, include_workings, overwrite_existing FROM scheduler_jobs WHERE active = 1'
        ).fetchall()

        for job in jobs:
            folder_path = job['folder_path']
            if not os.path.isdir(folder_path):
                connection.execute(
                    'UPDATE scheduler_jobs SET updated_at = ?, last_status = ?, last_message = ? WHERE job_id = ?',
                    (time.time(), 'error', 'Folder not found', job['job_id'])
                )
                continue

            for entry in os.listdir(folder_path):
                file_path = os.path.join(folder_path, entry)
                if not os.path.isfile(file_path):
                    continue
                if not re.search(r'FTP Input File \w+ \d{4}.*\.(xlsx|xls)$', entry, flags=re.IGNORECASE):
                    continue

                archive_path = f'{file_path}.processed'
                if os.path.exists(archive_path):
                    continue

                validation = _validate_upload_file(file_path, entry)
                if not validation['filename_valid'] or not validation['required_sheets_present']:
                    connection.execute(
                        'UPDATE scheduler_jobs SET updated_at = ?, last_status = ?, last_message = ? WHERE job_id = ?',
                        (time.time(), 'error', '; '.join(validation['warnings']) or 'Validation failed', job['job_id'])
                    )
                    continue

                scheduler_job_id = f'scheduler-{uuid.uuid4()}'
                temp_copy_path = f'{file_path}.{scheduler_job_id}.tmp'
                shutil.copyfile(file_path, temp_copy_path)
                _insert_upload_job(scheduler_job_id)
                _run_ftp_job(
                    scheduler_job_id,
                    temp_copy_path,
                    entry,
                    bool(job['include_workings']),
                    bool(job['overwrite_existing']),
                    export_warning='Processed automatically by scheduler.'
                )
                shutil.copyfile(file_path, archive_path)
                connection.execute(
                    'UPDATE scheduler_jobs SET updated_at = ?, last_run_at = ?, last_status = ?, last_message = ? WHERE job_id = ?',
                    (time.time(), time.time(), 'success', f'Processed {entry}', job['job_id'])
                )
                processed_files += 1
                _audit_event('scheduler_processed', details={'folder_path': folder_path, 'filename': entry})

        connection.commit()

    return processed_files


def _run_ftp_job(job_id, upload_path, filename, include_non_loan_sheets, overwrite_existing=False, export_warning=None):
    """Background thread: process workbook and write results."""
    global latest_data

    def log_stage(stage_name, t0):
        print(f"[PERF] {stage_name}: {perf_counter() - t0:.3f}s")

    def progress(pct, stage):
        _update_job(job_id, progress=pct, stage=stage)

    try:
        upload_start_time = perf_counter()
        progress(5, 'Preparing upload')

        filename_match = re.search(r'FTP Input File (\w+) (\d{4})', filename)
        if not filename_match:
            _update_job(job_id, status='error', error='Filename must be: FTP Input File Month Year.xlsx')
            return

        month_name = filename_match.group(1)
        year = int(filename_match.group(2))

        month_map = {
            'january':1,'february':2,'march':3,'april':4,'may':5,'june':6,
            'july':7,'august':8,'september':9,'october':10,'november':11,'december':12
        }
        month_num = month_map.get(month_name.lower())
        if not month_num:
            _update_job(job_id, status='error', error=f'Invalid month: {month_name}')
            return

        progress(8, f'Preparing upload for {month_name} {year}')

        report_key_suffix = None if overwrite_existing else _next_report_version_suffix(year, month_num)
        report_key = _make_report_key(year, month_num, report_key_suffix)
        workings_manifest_path = _build_workings_manifest_path(report_key)
        workings_manifest = {
            'report_key': report_key,
            'month': month_name,
            'year': year,
            'generated_at': time.time(),
            'sheets': {}
        }

        first_day = datetime(year, month_num, 1)
        last_day = (datetime(year + 1, 1, 1) if month_num == 12 else datetime(year, month_num + 1, 1)) - timedelta(days=1)

        progress(10, 'Reading workbook structure')
        read_start = perf_counter()
        excel_file = pd.ExcelFile(upload_path)
        sheet_names = excel_file.sheet_names
        log_stage('Read workbook metadata', read_start)
        print(f"Found sheets: {sheet_names}")

        output_stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_filename = f"FTP_Results_{month_name}_{year}_{output_stamp}.xlsx"
        excel_output_path = os.path.join(PROCESSED_OUTPUTS_DIR, excel_filename)
        loan_results_output_path = excel_output_path
        if include_non_loan_sheets:
            loan_results_output_path = os.path.join(
                PROCESSED_OUTPUTS_DIR,
                f"FTP_Results_LoansOnly_{month_name}_{year}_{output_stamp}.xlsx"
            )
        original_extension = os.path.splitext(filename)[1] or '.xlsx'
        original_upload_filename = f"FTP_Original_{month_name}_{year}_{output_stamp}{original_extension}"
        original_upload_path = os.path.join(PROCESSED_OUTPUTS_DIR, original_upload_filename)

        try:
            shutil.copyfile(upload_path, original_upload_path)
        except OSError as exc:
            original_upload_path = None
            original_upload_filename = None
            print(f"[WARN] Failed to archive original upload for {filename}: {exc}")

        sheets_data = {}
        global_summaries = {'ZWG': {}, 'FX': {}}
        skipped_sheet_count = 0

        branch_sbu_map = load_branch_sbu_map()
        branch_sbu_lookup = {code: v.get('sbu', 'Unknown') for code, v in branch_sbu_map.items()}

        loan_sheet_names = [s for s in sheet_names if s in LOAN_SHEETS]
        total_loan_sheets = max(len(loan_sheet_names), 1)

        excel_write_start = perf_counter()
        if include_non_loan_sheets:
            # openpyxl is significantly more memory-hungry on large workbooks.
            # Use xlsxwriter without constant_memory for full-workbook exports.
            writer_engine = 'xlsxwriter'
            writer_kwargs = {}
        else:
            writer_engine = 'xlsxwriter'
            writer_kwargs = {'engine_kwargs': {'options': {'constant_memory': True}}}

        def process_sheets(writer):
            nonlocal skipped_sheet_count
            completed_loan = 0

            for sheet in sheet_names:
                print(f"Processing: {sheet}")
                sheet_start = perf_counter()

                if sheet not in LOAN_SHEETS and not include_non_loan_sheets:
                    skipped_sheet_count += 1
                    sheets_data[sheet] = {'columns': [], 'data': [], 'shape': [0, 0],
                                          'note': 'Skipped (memory optimisation)'}
                    log_stage(f'Sheet {sheet} (skipped)', sheet_start)
                    continue

                if sheet not in LOAN_SHEETS and include_non_loan_sheets:
                    # Keep non-loan workings in-place from original workbook and avoid pandas parse overhead.
                    sheets_data[sheet] = {
                        'columns': [],
                        'data': [],
                        'shape': [0, 0],
                        'note': 'Included from original workbook via low-memory merge'
                    }
                    log_stage(f'Sheet {sheet} (kept from original)', sheet_start)
                    continue

                if sheet in LOAN_SHEETS:
                    branch_column_candidates = ['Branch Code', 'BRANCHCODE', 'BRANCH_CODE']

                    # FTP computation constants
                    first_day_ts = pd.Timestamp(first_day.date())
                    last_day_ts = pd.Timestamp(last_day.date())
                    full_period = (last_day_ts - first_day_ts).days + 1
                    bucket_labels = ['<7days','7-14days','14-21days','21-30days','30-60days','60-90days',
                                     '90-180days','180-270days','270-360days','360-720days',
                                     '720-1080days','1080-1460days','1460-1800days','+1800days']
                    bin_edges = [0, 7, 14, 21, 30, 60, 90, 180, 270, 360, 720, 1080, 1460, 1800, float('inf')]
                    rates = zwg_rates if sheet == 'ZWG LOANS' else usd_rates
                    rv = np.array([(rates[i] if i < len(rates) else rates[-1]) / 100
                                   for i in range(len(bucket_labels))], dtype=np.float32)
                    currency = 'ZWG' if sheet == 'ZWG LOANS' else 'FX'

                    sheet_data_path = _build_workings_sheet_data_path(report_key, sheet)
                    raw_data_path = sheet_data_path.replace('.ndjson', '_raw.ndjson')

                    raw_stream_result = _stream_workbook_sheet_to_ndjson(upload_path, sheet, raw_data_path)
                    if raw_stream_result is None:
                        progress(18 + int(40 * completed_loan / total_loan_sheets), f'Reading raw rows from {sheet}')
                        try:
                            df_raw = excel_file.parse(sheet_name=sheet)
                        except Exception as exc:
                            print(f"Error parsing {sheet}: {exc}")
                            continue

                        total_rows = len(df_raw)
                        print(f"[WRITE_RAW] {sheet}: Writing {total_rows} raw rows to {raw_data_path}")
                        with open(raw_data_path, 'w', encoding='utf-8') as raw_file:
                            for _, row in df_raw.iterrows():
                                row_dict = {col: _sanitize_json_value(row.get(col)) for col in df_raw.columns}
                                raw_file.write(json.dumps(row_dict, ensure_ascii=True, allow_nan=False, default=_json_default))
                                raw_file.write('\n')
                        del df_raw
                        gc.collect()
                        print(f"[WRITE_RAW] {sheet}: Raw data written, DataFrame freed")
                    else:
                        total_rows = int(raw_stream_result.get('row_count') or 0)
                        print(f"[WRITE_RAW] {sheet}: Streamed {total_rows} raw rows directly to {raw_data_path}")

                    # STEP 2: Process chunks from raw JSON, write computed JSON
                    chunk_size = 50000
                    summary_exposure = 0.0
                    summary_ftp = 0.0
                    summary_by_sbu = {}
                    preview_rows = []
                    processed_columns = None
                    sheet_row_cursor = 0

                    print(f"[PROCESS] {sheet}: Processing {total_rows} rows in chunks of {chunk_size}")
                    chunk_buffer = []
                    chunk_count = 0
                    output_file = open(sheet_data_path, 'w', encoding='utf-8')
                    progress(20 + int(45 * completed_loan / total_loan_sheets), f'Computing FTP for {sheet}')

                    try:
                        with open(raw_data_path, 'r', encoding='utf-8') as raw_file:
                            for line_idx, line in enumerate(raw_file):
                                if not line.strip():
                                    continue
                                raw_row = json.loads(line)
                                chunk_buffer.append(raw_row)

                                # Process chunk when full or at EOF
                                if len(chunk_buffer) >= chunk_size or line_idx == total_rows - 1:
                                    chunk = pd.DataFrame(chunk_buffer)
                                    chunk_idx = line_idx - len(chunk_buffer) + 1
                                    chunk_end = min(line_idx + 1, total_rows)
                                    print(f"[CHUNK] {sheet} rows {chunk_idx}-{chunk_end}/{total_rows}")
                                    chunk_pct = 20 + int(45 * (completed_loan + (chunk_end / max(total_rows, 1))) / total_loan_sheets)
                                    progress(chunk_pct, f'Computing FTP for {month_name} {year}: {sheet} {chunk_end}/{total_rows}')

                                    # Add SBU mapping
                                    branch_col = next((column for column in branch_column_candidates if column in chunk.columns), None)
                                    if branch_col:
                                        chunk[branch_col] = chunk[branch_col].astype(str).str.strip()
                                        chunk['SBU'] = chunk[branch_col].map(branch_sbu_lookup).fillna('Unknown')
                                    else:
                                        chunk['SBU'] = 'Unknown'

                                    # Parse and convert dates
                                    if 'BOOKING_DATE' in chunk.columns:
                                        chunk['BOOKING_DATE'] = pd.to_datetime(chunk['BOOKING_DATE'], errors='coerce').fillna(first_day)
                                    else:
                                        chunk['BOOKING_DATE'] = first_day
                                    if 'MATURITY_DATE' in chunk.columns:
                                        chunk['MATURITY_DATE'] = pd.to_datetime(chunk['MATURITY_DATE'], errors='coerce').fillna(first_day + timedelta(days=365))
                                    else:
                                        chunk['MATURITY_DATE'] = first_day + timedelta(days=365)

                                    # Compute TENOR
                                    chunk['TENOR'] = (chunk['MATURITY_DATE'] - chunk['BOOKING_DATE']).dt.days.clip(lower=0)

                                    # Compute DimDays, DTM, MTM
                                    bd = chunk['BOOKING_DATE']
                                    md = chunk['MATURITY_DATE']
                                    chunk['DimDays'] = np.where(
                                        (bd <= first_day_ts) & (md >= last_day_ts), full_period,
                                        np.where((bd >= first_day_ts) & (md >= last_day_ts), (last_day_ts - bd).dt.days + 1,
                                        np.where((bd >= first_day_ts) & (md <= last_day_ts), (md - bd).dt.days,
                                                 (md - first_day_ts).dt.days))
                                    ).astype(np.int32)
                                    chunk['DTM'] = np.where(md > last_day_ts, (md - last_day_ts).dt.days, 0).astype(np.int32)
                                    chunk['MTM'] = (chunk['DTM'] / 30).round(1).astype(np.float32)

                                    # Compute buckets and FTP charges
                                    exposure = pd.to_numeric(
                                        chunk['Currency Exposure + Currency Accrued Reporting'], errors='coerce'
                                    ).fillna(0).astype(np.float32)
                                    mtm_days = (chunk['MTM'] * 30).astype(np.float32)
                                    bucket_idx = pd.cut(mtm_days, bins=bin_edges, labels=False, right=False, include_lowest=True)
                                    bucket_idx = bucket_idx.fillna(len(bucket_labels) - 1).astype(int)

                                    bv = np.zeros((len(chunk), len(bucket_labels)), dtype=np.float32)
                                    ev = exposure.to_numpy(dtype=np.float32, copy=False)
                                    iv = bucket_idx.to_numpy(dtype=np.int16, copy=False)
                                    pos = np.nonzero(ev > 0)[0]
                                    bv[pos, iv[pos]] = ev[pos]
                                    chunk[bucket_labels] = bv
                                    chunk['Check'] = bv.sum(axis=1).astype(np.float32)
                                    chunk['FTP Charge'] = (bv @ rv).astype(np.float32)

                                    # Accumulate summaries
                                    chunk_exposure = chunk['Currency Exposure + Currency Accrued Reporting'].sum()
                                    chunk_ftp = chunk['FTP Charge'].sum()
                                    summary_exposure += chunk_exposure
                                    summary_ftp += chunk_ftp

                                    chunk_sbu = chunk.groupby('SBU').agg({
                                        'Currency Exposure + Currency Accrued Reporting': 'sum',
                                        'FTP Charge': 'sum'
                                    }).reset_index()
                                    for _, row in chunk_sbu.iterrows():
                                        sbu_name = row['SBU']
                                        if sbu_name not in summary_by_sbu:
                                            summary_by_sbu[sbu_name] = {'Currency Exposure + Currency Accrued Reporting': 0.0, 'FTP Charge': 0.0, 'by_branch': {}}
                                        summary_by_sbu[sbu_name]['Currency Exposure + Currency Accrued Reporting'] += row['Currency Exposure + Currency Accrued Reporting']
                                        summary_by_sbu[sbu_name]['FTP Charge'] += row['FTP Charge']
                                    
                                    # Also accumulate branch-level details within each SBU
                                    if branch_col and branch_col in chunk.columns:
                                        chunk_branch = chunk.groupby([branch_col, 'SBU']).agg({
                                            'Currency Exposure + Currency Accrued Reporting': 'sum',
                                            'FTP Charge': 'sum'
                                        }).reset_index()
                                        for _, brow in chunk_branch.iterrows():
                                            branch_code = str(brow[branch_col]).strip()
                                            sbu_name = brow['SBU']
                                            exposure_val = brow['Currency Exposure + Currency Accrued Reporting']
                                            ftp_val = brow['FTP Charge']
                                            
                                            # Get unit name from branch map
                                            branch_info = branch_sbu_map.get(branch_code, {})
                                            unit_name = branch_info.get('unit', 'Unknown')
                                            
                                            # Initialize branch entry if needed
                                            if sbu_name not in summary_by_sbu:
                                                summary_by_sbu[sbu_name] = {'Currency Exposure + Currency Accrued Reporting': 0.0, 'FTP Charge': 0.0, 'by_branch': {}}
                                            if 'by_branch' not in summary_by_sbu[sbu_name]:
                                                summary_by_sbu[sbu_name]['by_branch'] = {}
                                            
                                            branch_key = f'{branch_code}|{unit_name}'
                                            if branch_key not in summary_by_sbu[sbu_name]['by_branch']:
                                                summary_by_sbu[sbu_name]['by_branch'][branch_key] = {
                                                    'branch_code': branch_code,
                                                    'unit': unit_name,
                                                    'sbu': sbu_name,
                                                    'Currency Exposure + Currency Accrued Reporting': 0.0,
                                                    'FTP Charge': 0.0
                                                }
                                            summary_by_sbu[sbu_name]['by_branch'][branch_key]['Currency Exposure + Currency Accrued Reporting'] += exposure_val
                                            summary_by_sbu[sbu_name]['by_branch'][branch_key]['FTP Charge'] += ftp_val

                                    # Capture preview rows
                                    if len(preview_rows) < 100:
                                       
                                        preview_chunk = chunk.head(100 - len(preview_rows)).copy()
                                        for col in preview_chunk.select_dtypes(include=['datetime64']).columns:
                                            preview_chunk[col] = preview_chunk[col].astype(str).replace('NaT', None)
                                        preview_rows.extend(preview_chunk.to_dict(orient='records'))

                                    # Write computed rows to output JSON
                                    for _, row in chunk.iterrows():
                                        row_dict = {col: _sanitize_json_value(row.get(col)) for col in chunk.columns}
                                        output_file.write(json.dumps(row_dict, ensure_ascii=True, allow_nan=False, default=_json_default))
                                        output_file.write('\n')

                                    if processed_columns is None:
                                        processed_columns = list(chunk.columns)

                                    header = sheet_row_cursor == 0
                                    chunk.to_excel(
                                        writer,
                                        sheet_name=sheet[:31],
                                        index=False,
                                        header=header,
                                        startrow=sheet_row_cursor,
                                    )
                                    sheet_row_cursor += len(chunk) + (1 if header else 0)

                                    # Clean up computation arrays
                                    del exposure, mtm_days, bucket_idx, bv, ev, iv, pos, bd, md, chunk
                                    chunk_buffer = []
                                    chunk_count += 1
                                    gc.collect()

                    finally:
                        output_file.close()

                    # STEP 3: Cleanup raw data file
                    try:
                        os.remove(raw_data_path)
                    except OSError:
                        pass

                    # Build summaries
                    sbu_summary = [{'SBU': sbu, **vals} for sbu, vals in summary_by_sbu.items()]
                    global_summaries[currency][sheet] = {
                        'total_exposure': float(summary_exposure),
                        'total_ftp_charge': float(summary_ftp),
                        'by_sbu': sbu_summary,
                        'row_count': total_rows
                    }

                    sheets_data[sheet] = {
                        'columns': processed_columns or [],
                        'data': preview_rows,
                        'shape': (total_rows, len(processed_columns or []))
                    }

                    workings_manifest['sheets'][sheet] = {
                        'columns': processed_columns or [],
                        'data_path': sheet_data_path,
                        'row_count': total_rows
                    }
                    gc.collect()

                    completed_loan += 1
                    pct = 15 + int(75 * completed_loan / total_loan_sheets)
                    progress(pct, f'Computed {sheet}')

                else:
                    try:
                        df = excel_file.parse(sheet_name=sheet)
                    except Exception as exc:
                        print(f"Error parsing {sheet}: {exc}")
                        continue

                    preview = df.head(100).copy()
                    sheets_data[sheet] = {
                        'columns': df.columns.tolist(),
                        'data': preview.to_dict(orient='records'),
                        'shape': df.shape
                    }
                    df.to_excel(writer, sheet_name=sheet[:31], index=False)
                    del df

                gc.collect()
                log_stage(f'Sheet {sheet}', sheet_start)
                print(f"Completed: {sheet}")

        try:
            with pd.ExcelWriter(loan_results_output_path, engine=writer_engine, **writer_kwargs) as writer:
                process_sheets(writer)
        except (ImportError, ModuleNotFoundError, ValueError):
            with pd.ExcelWriter(loan_results_output_path, engine='openpyxl') as writer:
                process_sheets(writer)

        if include_non_loan_sheets:
            progress(90, 'Building workbook with workings sheets')
            try:
                _merge_processed_loans_into_original_workbook(upload_path, loan_results_output_path, excel_output_path)
            finally:
                if loan_results_output_path != excel_output_path and os.path.exists(loan_results_output_path):
                    try:
                        os.remove(loan_results_output_path)
                    except OSError as exc:
                        print(f"[WARN] Failed to remove temporary loan-only workbook {loan_results_output_path}: {exc}")

        log_stage('Write processed workbook', excel_write_start)
        progress(92, 'Saving results to archive')

        with open(workings_manifest_path, 'w', encoding='utf-8') as workings_manifest_file:
            json.dump(workings_manifest, workings_manifest_file, ensure_ascii=True, default=_json_default, allow_nan=False)

        snapshot_start_time = perf_counter()
        latest_data['filename'] = filename
        latest_data['sheets'] = sheets_data
        latest_data['summaries'] = global_summaries
        latest_data['period'] = {
            'first_day': first_day.strftime('%d %B %Y'),
            'last_day': last_day.strftime('%d %B %Y'),
            'month': month_name,
            'month_number': month_num,
            'year': year,
            'report_key': report_key,
            'version_suffix': report_key_suffix
        }
        latest_data['excel_output_path'] = excel_output_path
        latest_data['excel_filename'] = excel_filename
        latest_data['original_upload_path'] = original_upload_path
        latest_data['original_upload_filename'] = original_upload_filename
        latest_data['workings_manifest_path'] = workings_manifest_path
        latest_data['excel_contains_workings'] = include_non_loan_sheets
        latest_data['skipped_sheet_count'] = skipped_sheet_count
        latest_data['export_warning'] = export_warning

        save_latest_data_snapshot()
        log_stage('Persist snapshot', snapshot_start_time)
        log_stage('Total upload pipeline', upload_start_time)
        print(f"✅ Stored summaries: {list(global_summaries.keys())}")

        _update_job(job_id,
                    status='done',
                    progress=100,
                    stage='Complete',
                    result={
                        'status': 'success',
                        'summary': global_summaries,
                        'period': latest_data['period'],
                        'excel_contains_workings': include_non_loan_sheets,
                        'skipped_sheet_count': skipped_sheet_count,
                        'export_warning': export_warning,
                        'original_upload_available': bool(original_upload_path),
                        'original_upload_filename': original_upload_filename,
                        'workings_available': bool(workings_manifest.get('sheets')),
                        'report_key': report_key
                    })

    except Exception as exc:
        import traceback
        traceback.print_exc()
        _update_job(job_id, status='error', error=str(exc))
    finally:
        try:
            if upload_path and os.path.exists(upload_path):
                os.remove(upload_path)
        except OSError as exc:
            print(f"[WARN] Failed to remove temp upload {upload_path}: {exc}")


@app.route('/')
def index():
    from flask import send_from_directory
    return send_from_directory('.', 'index.html')


@app.route('/upload', methods=['POST'])
def upload_file():
    _cleanup_expired_jobs()

    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400

    filename = file.filename

    filename_match = re.search(r'FTP Input File (\w+) (\d{4})', filename)
    if not filename_match:
        return jsonify({'error': 'Filename must be: FTP Input File Month Year.xlsx'}), 400

    temp_upload_path = os.path.join(TEMP_UPLOADS_DIR, f'{uuid.uuid4()}_{os.path.basename(filename)}')
    try:
        file.save(temp_upload_path)
    except Exception as exc:
        return jsonify({'error': f'Failed to store upload temporarily: {str(exc)}'}), 500

    include_workings_raw = str(request.form.get('include_workings', '1')).strip().lower()
    include_workings_requested = include_workings_raw in {'1', 'true', 'yes', 'on'}
    force_full_workbook_raw = str(request.form.get('force_full_workbook', '0')).strip().lower()
    force_full_workbook = force_full_workbook_raw in {'1', 'true', 'yes', 'on'}
    include_non_loan_sheets = INCLUDE_NON_LOAN_SHEETS or include_workings_requested
    overwrite_existing_raw = str(request.form.get('overwrite_existing', '0')).strip().lower()
    overwrite_existing = overwrite_existing_raw in {'1', 'true', 'yes', 'on'}

    upload_size_mb = os.path.getsize(temp_upload_path) / (1024 * 1024)
    export_warning = None
    if include_non_loan_sheets and FORCE_RESULTS_ONLY_ON_HOSTED:
        include_non_loan_sheets = False
        export_warning = (
            'Full workbook export is disabled on this hosted deployment to prevent worker restarts. '
            'Use the processed Excel results, the original uploaded workbook, and the workings workbook generated from JSON.'
        )
    elif (
        include_non_loan_sheets
        and not force_full_workbook
        and not INCLUDE_NON_LOAN_SHEETS
        and upload_size_mb > INCLUDE_WORKINGS_MAX_UPLOAD_MB
    ):
        include_non_loan_sheets = False
        export_warning = (
            f'Include workings sheets was requested but automatically downgraded to results-only '
            f'because the upload size ({upload_size_mb:.1f} MB) exceeds the safe limit '
            f'({INCLUDE_WORKINGS_MAX_UPLOAD_MB:.1f} MB) for this hosting memory profile.'
        )
    elif include_non_loan_sheets and force_full_workbook and upload_size_mb > INCLUDE_WORKINGS_MAX_UPLOAD_MB:
        export_warning = (
            f'Force full workbook is enabled for an upload of {upload_size_mb:.1f} MB, which is above '
            f'the safe limit ({INCLUDE_WORKINGS_MAX_UPLOAD_MB:.1f} MB). Processing may restart under memory pressure.'
        )

    job_id = str(uuid.uuid4())
    now = time.time()
    with _jobs_lock:
        _jobs[job_id] = {
            'status': 'running',
            'progress': 0,
            'stage': 'Queued',
            'result': None,
            'error': None,
            'created_at': now,
            'updated_at': now,
            'completed_at': None
        }
    _insert_upload_job(job_id)

    thread = threading.Thread(
        target=_run_ftp_job,
        args=(job_id, temp_upload_path, filename, include_non_loan_sheets, overwrite_existing, export_warning),
        daemon=True
    )
    thread.start()

    return jsonify({'job_id': job_id, 'status': 'running'})


@app.route('/upload/preflight', methods=['POST'])
def upload_preflight():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400

    temp_upload_path = os.path.join(TEMP_UPLOADS_DIR, f'{uuid.uuid4()}_{os.path.basename(file.filename)}')
    try:
        file.save(temp_upload_path)
        validation = _validate_upload_file(temp_upload_path, file.filename)
        return jsonify(validation)
    except Exception as exc:
        return jsonify({'error': f'Failed to validate workbook: {str(exc)}'}), 500
    finally:
        if os.path.exists(temp_upload_path):
            try:
                os.remove(temp_upload_path)
            except OSError:
                pass


@app.route('/upload/status/<job_id>', methods=['GET'])
def upload_status(job_id):
    _cleanup_expired_jobs()

    # Use SQLite as the authoritative store so status checks work across workers.
    job = _fetch_upload_job(job_id)
    if not job:
        return jsonify({'error': 'Unknown job'}), 404
    response = {
        'status': job['status'],
        'progress': job['progress'],
        'stage': job['stage'],
    }
    if job['status'] == 'done':
        response['result'] = job['result']
    elif job['status'] == 'error':
        response['error'] = job['error']
    return jsonify(response)


# ---------------------------------------------------------------------------
# Legacy synchronous path kept for local dev / non-Render environments.
# Remove once async path is proven in production.
# ---------------------------------------------------------------------------
def _upload_file_sync():
    global latest_data

    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400

    temp_upload_path = os.path.join(TEMP_UPLOADS_DIR, f'{uuid.uuid4()}_{os.path.basename(file.filename)}')

    try:
        upload_start_time = perf_counter()

        def log_stage(stage_name, stage_start_time):
            elapsed_seconds = perf_counter() - stage_start_time
            print(f"[PERF] {stage_name}: {elapsed_seconds:.3f}s")

        # Parse filename
        filename_match = re.search(r'FTP Input File (\w+) (\d{4})', file.filename)
        if not filename_match:
            return jsonify({'error': 'Filename must be: FTP Input File Month Year.xlsx'}), 400
        
        month_name = filename_match.group(1)
        year = int(filename_match.group(2))
        
        month_map = {'january':1,'february':2,'march':3,'april':4,'may':5,'june':6,
                     'july':7,'august':8,'september':9,'october':10,'november':11,'december':12}
        month_num = month_map.get(month_name.lower())
        if not month_num:
            return jsonify({'error': f'Invalid month: {month_name}'}), 400

        report_key_suffix = None if overwrite_existing else _next_report_version_suffix(year, month_num)
        report_key = _make_report_key(year, month_num, report_key_suffix)

        try:
            file.save(temp_upload_path)
        except Exception as exc:
            return jsonify({'error': f'Failed to store upload temporarily: {str(exc)}'}), 500

        include_workings_raw = str(request.form.get('include_workings', '1')).strip().lower()
        include_workings_requested = include_workings_raw in {'1', 'true', 'yes', 'on'}
        include_non_loan_sheets = INCLUDE_NON_LOAN_SHEETS or include_workings_requested
        overwrite_existing_raw = str(request.form.get('overwrite_existing', '0')).strip().lower()
        overwrite_existing = overwrite_existing_raw in {'1', 'true', 'yes', 'on'}
        
        first_day = datetime(year, month_num, 1)
        if month_num == 12:
            last_day = datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            last_day = datetime(year, month_num + 1, 1) - timedelta(days=1)
        
        # Get sheet names
        read_start_time = perf_counter()
        excel_file = pd.ExcelFile(temp_upload_path)
        sheet_names = excel_file.sheet_names
        log_stage('Read workbook metadata', read_start_time)
        print(f"Found sheets: {sheet_names}")

        output_stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_filename = f"FTP_Results_{month_name}_{year}_{output_stamp}.xlsx"
        excel_output_path = os.path.join(PROCESSED_OUTPUTS_DIR, excel_filename)
        
        sheets_data = {}
        global_summaries = {'ZWG': {}, 'FX': {}}
        skipped_sheet_count = 0
        
        branch_sbu_map = load_branch_sbu_map()
        branch_sbu_lookup = {code: value.get('sbu', 'Unknown') for code, value in branch_sbu_map.items()}
        
        excel_write_start_time = perf_counter()
        if include_non_loan_sheets:
            # openpyxl can exceed memory limits for large files; prefer xlsxwriter here.
            writer_engine = 'xlsxwriter'
            writer_kwargs = {}
        else:
            writer_engine = 'xlsxwriter'
            writer_kwargs = {'engine_kwargs': {'options': {'constant_memory': True}}}
        try:
            with pd.ExcelWriter(excel_output_path, engine=writer_engine, **writer_kwargs) as writer:
                for sheet in sheet_names:
                    print(f"Processing: {sheet}")
                    sheet_start_time = perf_counter()

                    if sheet not in LOAN_SHEETS and not include_non_loan_sheets:
                        skipped_sheet_count += 1
                        sheets_data[sheet] = {
                            'columns': [],
                            'data': [],
                            'shape': [0, 0],
                            'note': 'Skipped for memory optimization (set INCLUDE_NON_LOAN_SHEETS=1 to include)'
                        }
                        log_stage(f'Sheet {sheet} (skipped)', sheet_start_time)
                        print(f"Completed: {sheet} (skipped)")
                        continue

                    try:
                        df = excel_file.parse(sheet_name=sheet)
                    except Exception as e:
                        print(f"Error: {e}")
                        continue

                    if sheet in LOAN_SHEETS:
                        df_processed = df
                        del df

                        # Add SBU column
                        branch_col = None
                        for col in ['Branch Code', 'BRANCHCODE', 'BRANCH_CODE']:
                            if col in df_processed.columns:
                                branch_col = col
                                break

                        if branch_col:
                            df_processed[branch_col] = df_processed[branch_col].astype(str).str.strip()
                            df_processed['SBU'] = df_processed[branch_col].map(branch_sbu_lookup).fillna('Unknown')
                        else:
                            df_processed['SBU'] = 'Unknown'

                        # Date processing
                        if 'BOOKING_DATE' in df_processed.columns:
                            df_processed['BOOKING_DATE'] = pd.to_datetime(df_processed['BOOKING_DATE'], errors='coerce')
                            df_processed['BOOKING_DATE'] = df_processed['BOOKING_DATE'].fillna(first_day)

                        if 'MATURITY_DATE' in df_processed.columns:
                            df_processed['MATURITY_DATE'] = pd.to_datetime(df_processed['MATURITY_DATE'], errors='coerce')
                            df_processed['MATURITY_DATE'] = df_processed['MATURITY_DATE'].fillna(first_day + timedelta(days=365))

                        # Calculate TENOR
                        if 'BOOKING_DATE' in df_processed.columns and 'MATURITY_DATE' in df_processed.columns:
                            df_processed['TENOR'] = (df_processed['MATURITY_DATE'] - df_processed['BOOKING_DATE']).dt.days
                            df_processed.loc[df_processed['TENOR'] < 0, 'TENOR'] = 0

                        # Calculate DimDays with vectorized date logic.
                        first_day_ts = pd.Timestamp(first_day.date())
                        last_day_ts = pd.Timestamp(last_day.date())
                        booking_dates = df_processed['BOOKING_DATE']
                        maturity_dates = df_processed['MATURITY_DATE']

                        full_period_days = (last_day_ts - first_day_ts).days + 1
                        dim_days = np.where(
                            (booking_dates <= first_day_ts) & (maturity_dates >= last_day_ts),
                            full_period_days,
                            np.where(
                                (booking_dates >= first_day_ts) & (maturity_dates >= last_day_ts),
                                (last_day_ts - booking_dates).dt.days + 1,
                                np.where(
                                    (booking_dates >= first_day_ts) & (maturity_dates <= last_day_ts),
                                    (maturity_dates - booking_dates).dt.days,
                                    (maturity_dates - first_day_ts).dt.days
                                )
                            )
                        )
                        df_processed['DimDays'] = dim_days.astype(np.int32)

                        # Calculate DTM and MTM
                        df_processed['DTM'] = np.where(
                            maturity_dates > last_day_ts,
                            (maturity_dates - last_day_ts).dt.days,
                            0
                        ).astype(np.int32)
                        df_processed['MTM'] = (df_processed['DTM'] / 30).round(1).astype(np.float32)

                        # Bucket columns
                        bucket_labels = ['<7days', '7-14days', '14-21days', '21-30days', '30-60days', '60-90days', '90-180days', '180-270days', '270-360days', '360-720days', '720-1080days', '1080-1460days', '1460-1800days', '+1800days']

                        # Allocate to buckets
                        exposure = pd.to_numeric(
                            df_processed['Currency Exposure + Currency Accrued Reporting'],
                            errors='coerce'
                        ).fillna(0).astype(np.float32)
                        mtm_days = (df_processed['MTM'] * 30).astype(np.float32)
                        tenors_list = [7,14,21,30,60,90,180,270,360,720,1080,1460,1800]
                        bin_edges = [0] + tenors_list + [float('inf')]
                        bucket_indices = pd.cut(mtm_days, bins=bin_edges, labels=False, right=False, include_lowest=True)
                        bucket_indices = bucket_indices.fillna(len(bucket_labels)-1).astype(int)

                        bucket_values = np.zeros((len(df_processed), len(bucket_labels)), dtype=np.float32)
                        exposure_values = exposure.to_numpy(dtype=np.float32, copy=False)
                        index_values = bucket_indices.to_numpy(dtype=np.int16, copy=False)
                        positive_exposure_rows = np.nonzero(exposure_values > 0)[0]
                        bucket_values[positive_exposure_rows, index_values[positive_exposure_rows]] = exposure_values[positive_exposure_rows]
                        df_processed[bucket_labels] = bucket_values

                        df_processed['Check'] = df_processed[bucket_labels].sum(axis=1).astype(np.float32)

                        # FTP Charge
                        rates = zwg_rates if sheet == "ZWG LOANS" else usd_rates
                        rate_vector = np.array(
                            [(rates[i] if i < len(rates) else rates[-1]) / 100 for i in range(len(bucket_labels))],
                            dtype=np.float32
                        )
                        df_processed['FTP Charge'] = (
                            df_processed[bucket_labels].to_numpy(dtype=np.float32, copy=False) @ rate_vector
                        ).astype(np.float32)

                        # Summary
                        currency = 'ZWG' if sheet == "ZWG LOANS" else 'FX'
                        sbu_summary = df_processed.groupby('SBU').agg({
                            'Currency Exposure + Currency Accrued Reporting': 'sum',
                            'FTP Charge': 'sum'
                        }).reset_index()

                        global_summaries[currency][sheet] = {
                            'total_exposure': float(df_processed['Currency Exposure + Currency Accrued Reporting'].sum()),
                            'total_ftp_charge': float(df_processed['FTP Charge'].sum()),
                            'by_sbu': sbu_summary.to_dict(orient='records'),
                            'row_count': len(df_processed)
                        }

                        # Preview
                        preview = df_processed.head(100).copy()
                        for col in preview.select_dtypes(include=['datetime64']).columns:
                            preview[col] = preview[col].astype(str).replace('NaT', None)

                        sheets_data[sheet] = {
                            'columns': df_processed.columns.tolist(),
                            'data': preview.to_dict(orient='records'),
                            'shape': df_processed.shape
                        }

                        # Persist full processed worksheet for Excel download
                        df_processed.to_excel(writer, sheet_name=sheet[:31], index=False)
                        del df_processed

                    else:
                        preview = df.head(100).copy()
                        sheets_data[sheet] = {
                            'columns': df.columns.tolist(),
                            'data': preview.to_dict(orient='records'),
                            'shape': df.shape
                        }
                        # Optional: include non-loan sheets in output when explicitly enabled.
                        df.to_excel(writer, sheet_name=sheet[:31], index=False)
                        del df

                    gc.collect()
                    log_stage(f'Sheet {sheet}', sheet_start_time)
                    print(f"Completed: {sheet}")
        except (ImportError, ModuleNotFoundError, ValueError):
            # Fallback when xlsxwriter is unavailable.
            with pd.ExcelWriter(excel_output_path, engine='openpyxl') as writer:
                for sheet in sheet_names:
                    print(f"Processing: {sheet}")
                    sheet_start_time = perf_counter()

                    if sheet not in LOAN_SHEETS and not include_non_loan_sheets:
                        skipped_sheet_count += 1
                        sheets_data[sheet] = {
                            'columns': [],
                            'data': [],
                            'shape': [0, 0],
                            'note': 'Skipped for memory optimization (set INCLUDE_NON_LOAN_SHEETS=1 to include)'
                        }
                        log_stage(f'Sheet {sheet} (skipped)', sheet_start_time)
                        print(f"Completed: {sheet} (skipped)")
                        continue

                    try:
                        df = excel_file.parse(sheet_name=sheet)
                    except Exception as e:
                        print(f"Error: {e}")
                        continue

                    if sheet in LOAN_SHEETS:
                        df_processed = df
                        del df

                        # Add SBU column
                        branch_col = None
                        for col in ['Branch Code', 'BRANCHCODE', 'BRANCH_CODE']:
                            if col in df_processed.columns:
                                branch_col = col
                                break

                        if branch_col:
                            df_processed[branch_col] = df_processed[branch_col].astype(str).str.strip()
                            df_processed['SBU'] = df_processed[branch_col].map(branch_sbu_lookup).fillna('Unknown')
                        else:
                            df_processed['SBU'] = 'Unknown'

                        # Date processing
                        if 'BOOKING_DATE' in df_processed.columns:
                            df_processed['BOOKING_DATE'] = pd.to_datetime(df_processed['BOOKING_DATE'], errors='coerce')
                            df_processed['BOOKING_DATE'] = df_processed['BOOKING_DATE'].fillna(first_day)

                        if 'MATURITY_DATE' in df_processed.columns:
                            df_processed['MATURITY_DATE'] = pd.to_datetime(df_processed['MATURITY_DATE'], errors='coerce')
                            df_processed['MATURITY_DATE'] = df_processed['MATURITY_DATE'].fillna(first_day + timedelta(days=365))

                        # Calculate TENOR
                        if 'BOOKING_DATE' in df_processed.columns and 'MATURITY_DATE' in df_processed.columns:
                            df_processed['TENOR'] = (df_processed['MATURITY_DATE'] - df_processed['BOOKING_DATE']).dt.days
                            df_processed.loc[df_processed['TENOR'] < 0, 'TENOR'] = 0

                        # Calculate DimDays with vectorized date logic.
                        first_day_ts = pd.Timestamp(first_day.date())
                        last_day_ts = pd.Timestamp(last_day.date())
                        booking_dates = df_processed['BOOKING_DATE']
                        maturity_dates = df_processed['MATURITY_DATE']

                        full_period_days = (last_day_ts - first_day_ts).days + 1
                        dim_days = np.where(
                            (booking_dates <= first_day_ts) & (maturity_dates >= last_day_ts),
                            full_period_days,
                            np.where(
                                (booking_dates >= first_day_ts) & (maturity_dates >= last_day_ts),
                                (last_day_ts - booking_dates).dt.days + 1,
                                np.where(
                                    (booking_dates >= first_day_ts) & (maturity_dates <= last_day_ts),
                                    (maturity_dates - booking_dates).dt.days,
                                    (maturity_dates - first_day_ts).dt.days
                                )
                            )
                        )
                        df_processed['DimDays'] = dim_days.astype(np.int32)

                        # Calculate DTM and MTM
                        df_processed['DTM'] = np.where(
                            maturity_dates > last_day_ts,
                            (maturity_dates - last_day_ts).dt.days,
                            0
                        ).astype(np.int32)
                        df_processed['MTM'] = (df_processed['DTM'] / 30).round(1).astype(np.float32)

                        # Bucket columns
                        bucket_labels = ['<7days', '7-14days', '14-21days', '21-30days', '30-60days', '60-90days', '90-180days', '180-270days', '270-360days', '360-720days', '720-1080days', '1080-1460days', '1460-1800days', '+1800days']

                        # Allocate to buckets
                        exposure = pd.to_numeric(
                            df_processed['Currency Exposure + Currency Accrued Reporting'],
                            errors='coerce'
                        ).fillna(0).astype(np.float32)
                        mtm_days = (df_processed['MTM'] * 30).astype(np.float32)
                        tenors_list = [7,14,21,30,60,90,180,270,360,720,1080,1460,1800]
                        bin_edges = [0] + tenors_list + [float('inf')]
                        bucket_indices = pd.cut(mtm_days, bins=bin_edges, labels=False, right=False, include_lowest=True)
                        bucket_indices = bucket_indices.fillna(len(bucket_labels)-1).astype(int)

                        bucket_values = np.zeros((len(df_processed), len(bucket_labels)), dtype=np.float32)
                        exposure_values = exposure.to_numpy(dtype=np.float32, copy=False)
                        index_values = bucket_indices.to_numpy(dtype=np.int16, copy=False)
                        positive_exposure_rows = np.nonzero(exposure_values > 0)[0]
                        bucket_values[positive_exposure_rows, index_values[positive_exposure_rows]] = exposure_values[positive_exposure_rows]
                        df_processed[bucket_labels] = bucket_values

                        df_processed['Check'] = df_processed[bucket_labels].sum(axis=1).astype(np.float32)

                        # FTP Charge
                        rates = zwg_rates if sheet == "ZWG LOANS" else usd_rates
                        rate_vector = np.array(
                            [(rates[i] if i < len(rates) else rates[-1]) / 100 for i in range(len(bucket_labels))],
                            dtype=np.float32
                        )
                        df_processed['FTP Charge'] = (
                            df_processed[bucket_labels].to_numpy(dtype=np.float32, copy=False) @ rate_vector
                        ).astype(np.float32)

                        # Summary
                        currency = 'ZWG' if sheet == "ZWG LOANS" else 'FX'
                        sbu_summary = df_processed.groupby('SBU').agg({
                            'Currency Exposure + Currency Accrued Reporting': 'sum',
                            'FTP Charge': 'sum'
                        }).reset_index()

                        global_summaries[currency][sheet] = {
                            'total_exposure': float(df_processed['Currency Exposure + Currency Accrued Reporting'].sum()),
                            'total_ftp_charge': float(df_processed['FTP Charge'].sum()),
                            'by_sbu': sbu_summary.to_dict(orient='records'),
                            'row_count': len(df_processed)
                        }

                        # Preview
                        preview = df_processed.head(100).copy()
                        for col in preview.select_dtypes(include=['datetime64']).columns:
                            preview[col] = preview[col].astype(str).replace('NaT', None)

                        sheets_data[sheet] = {
                            'columns': df_processed.columns.tolist(),
                            'data': preview.to_dict(orient='records'),
                            'shape': df_processed.shape
                        }

                        # Persist full processed worksheet for Excel download
                        df_processed.to_excel(writer, sheet_name=sheet[:31], index=False)
                        del df_processed

                    else:
                        preview = df.head(100).copy()
                        sheets_data[sheet] = {
                            'columns': df.columns.tolist(),
                            'data': preview.to_dict(orient='records'),
                            'shape': df.shape
                        }
                        # Optional: include non-loan sheets in output when explicitly enabled.
                        df.to_excel(writer, sheet_name=sheet[:31], index=False)
                        del df

                    gc.collect()
                    log_stage(f'Sheet {sheet}', sheet_start_time)
                    print(f"Completed: {sheet}")
            log_stage('Write processed workbook', excel_write_start_time)

            # Add SBU Summary Sheet to the Excel file
            try:
                from openpyxl.styles import Font, PatternFill, Alignment
                workbook = openpyxl.load_workbook(excel_output_path)
                
                # Remove 'Summary' sheet if it exists
                if 'Summary' in workbook.sheetnames:
                    del workbook['Summary']
                
                # Create Summary sheet at the beginning
                summary_sheet = workbook.create_sheet('Summary', 0)
                
                # Write headers
                headers = ['Currency', 'SBU', 'Exposure', 'FTP Charge']
                header_fill = PatternFill(start_color='7F1D1D', end_color='7F1D1D', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True)
                
                for col_idx, header in enumerate(headers, 1):
                    cell = summary_sheet.cell(row=1, column=col_idx)
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                
                # Write data
                row_idx = 2
                for currency in sorted(global_summaries.keys()):
                    sbu_aggregates = {}
                    
                    # Aggregate SBU data from all sheets
                    for sheet_data in global_summaries[currency].values():
                        if sheet_data.get('by_sbu'):
                            for sbu in sheet_data['by_sbu']:
                                sbu_name = sbu.get('SBU', 'Unknown')
                                if sbu_name not in sbu_aggregates:
                                    sbu_aggregates[sbu_name] = {'exposure': 0.0, 'ftp_charge': 0.0}
                                sbu_aggregates[sbu_name]['exposure'] += sbu.get('Currency Exposure + Currency Accrued Reporting', 0)
                                sbu_aggregates[sbu_name]['ftp_charge'] += sbu.get('FTP Charge', 0)
                    
                    # Write SBU rows
                    for sbu_name in sorted(sbu_aggregates.keys()):
                        sbu_vals = sbu_aggregates[sbu_name]
                        summary_sheet.cell(row=row_idx, column=1).value = currency
                        summary_sheet.cell(row=row_idx, column=2).value = sbu_name
                        summary_sheet.cell(row=row_idx, column=3).value = float(sbu_vals['exposure'])
                        summary_sheet.cell(row=row_idx, column=4).value = float(sbu_vals['ftp_charge'])
                        
                        # Format numbers
                        summary_sheet.cell(row=row_idx, column=3).number_format = '#,##0.00'
                        summary_sheet.cell(row=row_idx, column=4).number_format = '#,##0.00'
                        
                        row_idx += 1
                
                # Adjust column widths
                summary_sheet.column_dimensions['A'].width = 12
                summary_sheet.column_dimensions['B'].width = 20
                summary_sheet.column_dimensions['C'].width = 18
                summary_sheet.column_dimensions['D'].width = 18
                
                workbook.save(excel_output_path)
                workbook.close()
                print(f"✅ Added Summary sheet to {excel_filename}")
            except Exception as summary_error:
                print(f"[WARN] Failed to add Summary sheet: {summary_error}")
                import traceback
                traceback.print_exc()

            # Store data
        latest_data['filename'] = file.filename
        latest_data['sheets'] = sheets_data
        latest_data['summaries'] = global_summaries
        latest_data['period'] = {
            'first_day': first_day.strftime('%d %B %Y'),
            'last_day': last_day.strftime('%d %B %Y'),
            'month': month_name,
            'month_number': month_num,
            'year': year,
            'report_key': report_key,
            'version_suffix': report_key_suffix
        }
        latest_data['excel_output_path'] = excel_output_path
        latest_data['excel_filename'] = excel_filename
        latest_data['excel_contains_workings'] = include_non_loan_sheets
        latest_data['skipped_sheet_count'] = skipped_sheet_count

        snapshot_start_time = perf_counter()
        save_latest_data_snapshot()
        log_stage('Persist snapshot', snapshot_start_time)
        log_stage('Total upload pipeline', upload_start_time)
        
        print(f"✅ Stored summaries: {list(global_summaries.keys())}")
        
        return jsonify({
            'status': 'success',
            'summary': global_summaries,
            'period': latest_data['period'],
            'excel_contains_workings': include_non_loan_sheets,
            'skipped_sheet_count': skipped_sheet_count,
            'report_key': report_key
        })
    
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        try:
            if os.path.exists(temp_upload_path):
                os.remove(temp_upload_path)
        except OSError as exc:
            print(f"[WARN] Failed to remove temp upload {temp_upload_path}: {exc}")

@app.route('/calculate', methods=['POST'])
def calculate():
    data = request.json
    result = compute_ftp_components(data.get('deposit'), data.get('loan'), data.get('tenure'))
    return jsonify(result)

@app.route('/ftp-curve-data', methods=['GET'])
def ftp_curve_data():
    return jsonify({
        'tenors': tenors,
        'zwg': {'name': 'ZWG FTP Curve', 'rates': zwg_rates, 'color': '#b33a3a', 'borderColor': '#921f1f'},
        'usd': {'name': 'USD FTP Curve', 'rates': usd_rates, 'color': '#2563eb', 'borderColor': '#1d4ed8'}
    })


@app.route('/ftp-curve-data', methods=['POST'])
def update_ftp_curve_data():
    global tenors, usd_rates, zwg_rates

    payload = request.get_json(silent=True) or {}
    updated_tenors = payload.get('tenors')
    updated_usd_rates = payload.get('usd_rates')
    updated_zwg_rates = payload.get('zwg_rates')

    if not isinstance(updated_tenors, list) or not isinstance(updated_usd_rates, list) or not isinstance(updated_zwg_rates, list):
        return jsonify({'error': 'tenors, usd_rates, and zwg_rates must all be arrays'}), 400

    if not updated_tenors or len(updated_tenors) != len(updated_usd_rates) or len(updated_tenors) != len(updated_zwg_rates):
        return jsonify({'error': 'Curve arrays must be non-empty and have matching lengths'}), 400

    try:
        normalized_tenors = [int(value) for value in updated_tenors]
        normalized_usd_rates = [float(value) for value in updated_usd_rates]
        normalized_zwg_rates = [float(value) for value in updated_zwg_rates]
    except (TypeError, ValueError):
        return jsonify({'error': 'All curve values must be numeric'}), 400

    if any(value <= 0 for value in normalized_tenors):
        return jsonify({'error': 'Tenors must be greater than zero'}), 400

    if normalized_tenors != sorted(normalized_tenors):
        return jsonify({'error': 'Tenors must be sorted in ascending order'}), 400

    curve_config = {
        'tenors': normalized_tenors,
        'usd_rates': normalized_usd_rates,
        'zwg_rates': normalized_zwg_rates
    }

    try:
        save_curve_config(curve_config)
    except OSError as exc:
        return jsonify({'error': f'Unable to save curve config: {str(exc)}'}), 500

    tenors = normalized_tenors
    usd_rates = normalized_usd_rates
    zwg_rates = normalized_zwg_rates

    return jsonify({
        'status': 'success',
        'tenors': tenors,
        'zwg': {'name': 'ZWG FTP Curve', 'rates': zwg_rates, 'color': '#b33a3a', 'borderColor': '#921f1f'},
        'usd': {'name': 'USD FTP Curve', 'rates': usd_rates, 'color': '#2563eb', 'borderColor': '#1d4ed8'}
    })


@app.route('/branch-sbu-map', methods=['GET'])
def get_branch_sbu_map():
    try:
        return jsonify({'status': 'success', 'mappings': load_branch_sbu_rows()})
    except Exception as e:
        return jsonify({'error': f'Failed to load branch mappings: {str(e)}'}), 500


@app.route('/branch-sbu-map', methods=['POST'])
def update_branch_sbu_map():
    payload = request.get_json(silent=True) or {}
    mappings = payload.get('mappings')

    if not isinstance(mappings, list):
        return jsonify({'error': 'mappings must be an array'}), 400

    try:
        save_branch_sbu_rows(mappings)
        return jsonify({'status': 'success', 'mappings': load_branch_sbu_rows()})
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': f'Failed to save branch mappings: {str(e)}'}), 500


@app.route('/branch-sbu-map/reset', methods=['POST'])
def reset_branch_sbu_map():
    try:
        reset_branch_sbu_map_to_default()
        return jsonify({'status': 'success', 'mappings': load_branch_sbu_rows()})
    except Exception as e:
        return jsonify({'error': f'Failed to reset branch mappings: {str(e)}'}), 500

@app.route('/get-preview', methods=['GET'])
def get_preview():
    report_key = request.args.get('report_key')
    month = request.args.get('month')
    year = request.args.get('year', type=int)
    if report_key:
        if load_latest_data_snapshot(report_key=report_key):
            return _strict_json_response({'filename': latest_data['filename'], 'sheets': latest_data['sheets'], 'summaries': latest_data.get('summaries', {}), 'period': latest_data.get('period')})
        return _strict_json_response({'message': 'No data found for that report version'}, status_code=404)
    if month and year:
        if load_latest_data_snapshot(month=month, year=year):
            return _strict_json_response({'filename': latest_data['filename'], 'sheets': latest_data['sheets'], 'summaries': latest_data.get('summaries', {}), 'period': latest_data.get('period')})
        return _strict_json_response({'message': 'No data found for that month and year'}, status_code=404)

    if latest_data['sheets'] or load_latest_data_snapshot():
        return _strict_json_response({'filename': latest_data['filename'], 'sheets': latest_data['sheets'], 'summaries': latest_data.get('summaries', {})})
    return _strict_json_response({'message': 'No data uploaded yet'}, status_code=404)


@app.route('/processed-reports', methods=['GET'])
def processed_reports():
    try:
        _ensure_processed_reports_db()
        month = request.args.get('month')
        year = request.args.get('year', type=int)
        with sqlite3.connect(PROCESSED_REPORTS_DB_PATH) as connection:
            connection.row_factory = sqlite3.Row
            if month and year:
                month_number = _normalize_month_number(month)
                rows = connection.execute(
                    '''
                    SELECT report_key, month_number, month_name, year, filename, excel_filename, updated_at, excel_contains_workings, skipped_sheet_count, export_warning, original_upload_path, original_upload_filename
                    FROM processed_reports
                    WHERE month_number = ? AND year = ?
                    ORDER BY updated_at DESC
                    ''',
                    (month_number, int(year))
                ).fetchall()
            else:
                rows = connection.execute(
                    '''
                    SELECT report_key, month_number, month_name, year, filename, excel_filename, updated_at, excel_contains_workings, skipped_sheet_count, export_warning, original_upload_path, original_upload_filename
                    FROM processed_reports
                    ORDER BY year DESC, month_number DESC, updated_at DESC
                    '''
                ).fetchall()

        return jsonify({
            'reports': [
                {
                    'report_key': row['report_key'],
                    'month': row['month_name'],
                    'month_number': row['month_number'],
                    'year': row['year'],
                    'filename': row['filename'],
                    'excel_filename': row['excel_filename'],
                    'updated_at': row['updated_at'],
                    'excel_contains_workings': bool(row['excel_contains_workings']),
                    'skipped_sheet_count': row['skipped_sheet_count'],
                    'export_warning': row['export_warning'],
                    'original_upload_available': bool(row['original_upload_path']) and os.path.exists(row['original_upload_path']),
                    'original_upload_filename': row['original_upload_filename'],
                    'workings_available': os.path.exists(_build_workings_manifest_path(row['report_key'])),
                    'notes_count': len(_load_report_notes(row['report_key']))
                }
                for row in rows
            ]
        })
    except Exception as exc:
        return jsonify({'error': f'Failed to list processed reports: {str(exc)}'}), 500


@app.route('/processed-reports/<report_key>/metadata', methods=['GET'])
def processed_report_metadata(report_key):
    try:
        row = _fetch_report_metadata(report_key)
        if not row:
            return jsonify({'error': 'Report version not found'}), 404

        summaries = json.loads(row['summaries_json']) if row['summaries_json'] else {}
        sheets = json.loads(row['sheets_json']) if row['sheets_json'] else {}
        total_rows = sum((sheet_data.get('row_count') or 0) for currency in summaries.values() for sheet_data in currency.values())

        return jsonify({
            'report_key': row['report_key'],
            'month': row['month_name'],
            'month_number': row['month_number'],
            'year': row['year'],
            'filename': row['filename'],
            'excel_filename': row['excel_filename'],
            'original_upload_filename': row['original_upload_filename'],
            'original_upload_available': bool(row['original_upload_path']) and os.path.exists(row['original_upload_path']),
            'workings_available': os.path.exists(_build_workings_manifest_path(row['report_key'])),
            'excel_contains_workings': bool(row['excel_contains_workings']),
            'skipped_sheet_count': row['skipped_sheet_count'],
            'export_warning': row['export_warning'],
            'sheet_count': len(sheets),
            'total_rows': total_rows,
            'file_size_bytes': os.path.getsize(row['excel_output_path']) if row['excel_output_path'] and os.path.exists(row['excel_output_path']) else None,
            'original_file_size_bytes': os.path.getsize(row['original_upload_path']) if row['original_upload_path'] and os.path.exists(row['original_upload_path']) else None,
            'notes': _load_report_notes(report_key),
        })
    except Exception as exc:
        return jsonify({'error': f'Failed to load report metadata: {str(exc)}'}), 500


@app.route('/processed-reports/<report_key>/notes', methods=['GET', 'POST'])
def processed_report_notes(report_key):
    try:
        if request.method == 'GET':
            return jsonify({'notes': _load_report_notes(report_key)})

        role_error = _require_role('admin', 'editor')
        if role_error:
            return role_error

        payload = request.get_json(silent=True) or {}
        note_text = str(payload.get('note', '')).strip()
        if not note_text:
            return jsonify({'error': 'note is required'}), 400

        _ensure_processed_reports_db()
        with sqlite3.connect(PROCESSED_REPORTS_DB_PATH) as connection:
            connection.execute(
                'INSERT INTO report_notes (report_key, note_text, created_by, created_at) VALUES (?, ?, ?, ?)',
                (report_key, note_text, _current_actor(), time.time())
            )
            connection.commit()

        _audit_event('note_added', report_key=report_key, details={'note': note_text})
        return jsonify({'status': 'success', 'notes': _load_report_notes(report_key)})
    except Exception as exc:
        return jsonify({'error': f'Failed to manage report notes: {str(exc)}'}), 500


@app.route('/processed-reports/compare', methods=['GET'])
def compare_processed_reports():
    report_key_a = request.args.get('report_key_a')
    report_key_b = request.args.get('report_key_b')
    if not report_key_a or not report_key_b:
        return jsonify({'error': 'report_key_a and report_key_b are required'}), 400

    row_a = _fetch_report_metadata(report_key_a)
    row_b = _fetch_report_metadata(report_key_b)
    if not row_a or not row_b:
        return jsonify({'error': 'One or both report versions were not found'}), 404

    summaries_a = json.loads(row_a['summaries_json']) if row_a['summaries_json'] else {}
    summaries_b = json.loads(row_b['summaries_json']) if row_b['summaries_json'] else {}

    comparison = []
    currencies = sorted(set(summaries_a.keys()) | set(summaries_b.keys()))
    for currency in currencies:
        sheets = sorted(set((summaries_a.get(currency) or {}).keys()) | set((summaries_b.get(currency) or {}).keys()))
        for sheet in sheets:
            left = (summaries_a.get(currency) or {}).get(sheet, {})
            right = (summaries_b.get(currency) or {}).get(sheet, {})
            comparison.append({
                'currency': currency,
                'sheet': sheet,
                'exposure_delta': float((right.get('total_exposure') or 0) - (left.get('total_exposure') or 0)),
                'ftp_delta': float((right.get('total_ftp_charge') or 0) - (left.get('total_ftp_charge') or 0)),
                'row_count_delta': int((right.get('row_count') or 0) - (left.get('row_count') or 0)),
            })

    _audit_event('report_compared', details={'report_key_a': report_key_a, 'report_key_b': report_key_b})
    return jsonify({'comparison': comparison})


@app.route('/audit-log', methods=['GET'])
def audit_log():
    try:
        limit = request.args.get('limit', default=100, type=int)
        with sqlite3.connect(PROCESSED_REPORTS_DB_PATH) as connection:
            connection.row_factory = sqlite3.Row
            rows = connection.execute(
                'SELECT event_id, event_type, report_key, actor, actor_role, details_json, created_at FROM audit_log ORDER BY created_at DESC LIMIT ?',
                (max(1, min(limit, 500)),)
            ).fetchall()

        return jsonify({
            'events': [
                {
                    'event_id': row['event_id'],
                    'event_type': row['event_type'],
                    'report_key': row['report_key'],
                    'actor': row['actor'],
                    'actor_role': row['actor_role'],
                    'details': json.loads(row['details_json']) if row['details_json'] else {},
                    'created_at': row['created_at'],
                }
                for row in rows
            ]
        })
    except Exception as exc:
        return jsonify({'error': f'Failed to load audit log: {str(exc)}'}), 500


@app.route('/notifications', methods=['GET'])
def notifications():
    try:
        with sqlite3.connect(PROCESSED_REPORTS_DB_PATH) as connection:
            connection.row_factory = sqlite3.Row
            rows = connection.execute(
                "SELECT event_id, report_key, details_json, created_at FROM audit_log WHERE event_type = 'notification' ORDER BY created_at DESC LIMIT 20"
            ).fetchall()

        return jsonify({
            'notifications': [
                {
                    'event_id': row['event_id'],
                    'report_key': row['report_key'],
                    'created_at': row['created_at'],
                    **(json.loads(row['details_json']) if row['details_json'] else {})
                }
                for row in rows
            ]
        })
    except Exception as exc:
        return jsonify({'error': f'Failed to load notifications: {str(exc)}'}), 500


@app.route('/settings/retention', methods=['GET'])
def retention_settings():
    return jsonify({
        'max_versions_per_month': REPORT_RETENTION_MAX_VERSIONS,
        'max_months': REPORT_RETENTION_MAX_MONTHS,
        'role_enforced': FTP_REQUIRE_ROLE,
        'default_role': FTP_DEFAULT_ROLE,
        'scheduler_enabled': ENABLE_SCHEDULER,
    })


@app.route('/scheduler-jobs', methods=['GET', 'POST'])
def scheduler_jobs():
    try:
        _ensure_processed_reports_db()
        if request.method == 'GET':
            with sqlite3.connect(PROCESSED_REPORTS_DB_PATH) as connection:
                connection.row_factory = sqlite3.Row
                rows = connection.execute(
                    'SELECT job_id, folder_path, include_workings, overwrite_existing, active, created_at, updated_at, last_run_at, last_status, last_message FROM scheduler_jobs ORDER BY updated_at DESC'
                ).fetchall()
            return jsonify({'jobs': [dict(row) for row in rows]})

        role_error = _require_role('admin')
        if role_error:
            return role_error

        payload = request.get_json(silent=True) or {}
        folder_path = str(payload.get('folder_path', '')).strip()
        if not folder_path:
            return jsonify({'error': 'folder_path is required'}), 400

        now_ts = time.time()
        with sqlite3.connect(PROCESSED_REPORTS_DB_PATH) as connection:
            connection.execute(
                '''
                INSERT INTO scheduler_jobs (folder_path, include_workings, overwrite_existing, active, created_at, updated_at)
                VALUES (?, ?, ?, 1, ?, ?)
                ''',
                (folder_path, int(bool(payload.get('include_workings'))), int(bool(payload.get('overwrite_existing'))), now_ts, now_ts)
            )
            connection.commit()

        _audit_event('scheduler_job_added', details={'folder_path': folder_path})
        return scheduler_jobs()
    except Exception as exc:
        return jsonify({'error': f'Failed to manage scheduler jobs: {str(exc)}'}), 500


@app.route('/scheduler-jobs/run', methods=['POST'])
def run_scheduler_jobs():
    try:
        role_error = _require_role('admin')
        if role_error:
            return role_error

        processed_files = _run_scheduler_cycle()
        return jsonify({'status': 'success', 'processed_files': processed_files})
    except Exception as exc:
        return jsonify({'error': f'Failed to run scheduler: {str(exc)}'}), 500


@app.route('/processed-reports/<report_key>', methods=['DELETE'])
def delete_processed_report_route(report_key):
    try:
        if not report_key:
            return jsonify({'error': 'report_key is required'}), 400

        deleted = delete_processed_report(report_key)
        if not deleted:
            return jsonify({'error': 'Report version not found'}), 404

        return jsonify({'status': 'success', 'report_key': report_key})
    except Exception as exc:
        return jsonify({'error': f'Failed to delete processed report: {str(exc)}'}), 500


_ensure_processed_reports_db()

if __name__ == '__main__':
    app.run(debug=True, port=5000)